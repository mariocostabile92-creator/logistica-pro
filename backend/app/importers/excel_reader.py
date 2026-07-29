import csv
import io
from pathlib import Path
from typing import Any

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
import xlrd

from app.core.config import ALLOWED_EXTENSIONS, ALLOWED_MIME_TYPES, DEFAULT_PREVIEW_ROWS, MAX_UPLOAD_SIZE_BYTES


def validate_upload(file: UploadFile, content: bytes) -> None:
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Formato file non supportato. Usa .xlsx, .xls o .csv.")
    if file.content_type and file.content_type not in ALLOWED_MIME_TYPES:
        raise HTTPException(status_code=400, detail="Tipo MIME non valido per il file caricato.")
    if len(content) > MAX_UPLOAD_SIZE_BYTES:
        raise HTTPException(status_code=413, detail="File troppo grande.")
    if not content:
        raise HTTPException(status_code=400, detail="File vuoto.")


def _clean_cell(value: Any) -> Any:
    if value is None:
        return ""
    return value


def _rows_to_dicts(rows: list[list[Any]], limit: int | None = None) -> tuple[list[str], list[dict[str, Any]]]:
    if not rows:
        return [], []
    headers = [str(_clean_cell(cell)).strip() or f"colonna_{idx + 1}" for idx, cell in enumerate(rows[0])]
    data_rows = rows[1 : limit + 1 if limit else None]
    records = []
    for row in data_rows:
        record = {header: _clean_cell(row[idx]) if idx < len(row) else "" for idx, header in enumerate(headers)}
        if any(str(value).strip() for value in record.values()):
            records.append(record)
    return headers, records


def read_tabular(content: bytes, filename: str, sheet_name: str | None = None, preview_only: bool = False) -> dict[str, Any]:
    suffix = Path(filename).suffix.lower()
    limit = DEFAULT_PREVIEW_ROWS if preview_only else None
    try:
        if suffix == ".csv":
            text = content.decode("utf-8-sig")
            sample = text[:2048]
            dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t") if sample else csv.excel
            rows = list(csv.reader(io.StringIO(text), dialect))
            columns, records = _rows_to_dicts(rows, limit)
            return {"sheets": [], "selected_sheet": None, "columns": columns, "rows": records}

        if suffix == ".xls":
            workbook = xlrd.open_workbook(file_contents=content)
            names = workbook.sheet_names()
            selected = sheet_name if sheet_name in names else names[0]
            sheet = workbook.sheet_by_name(selected)
            rows = [[sheet.cell_value(row_index, col_index) for col_index in range(sheet.ncols)] for row_index in range(sheet.nrows)]
            columns, records = _rows_to_dicts(rows, limit)
            return {"sheets": names, "selected_sheet": selected, "columns": columns, "rows": records}

        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        selected = sheet_name if sheet_name in workbook.sheetnames else workbook.sheetnames[0]
        sheet = workbook[selected]
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        columns, records = _rows_to_dicts(rows, limit)
        workbook.close()
        return {"sheets": workbook.sheetnames, "selected_sheet": selected, "columns": columns, "rows": records}
    except Exception as exc:
        raise HTTPException(status_code=400, detail="File non leggibile o corrotto.") from exc
