from datetime import date, datetime, time
from decimal import Decimal
from typing import Any

from openpyxl.utils import get_column_letter

from app.importers.workbook_profiler.errors import WorkbookSelectionError
from app.importers.workbook_profiler.models import (
    MappingFieldOption,
    ProfileIssue,
    ProfiledWorkbook,
    ScannedSheet,
    WorkbookType,
)
from app.importers.workbook_profiler.sheet_profiler import profile_sheets
from app.importers.workbook_profiler.workbook_classifier import (
    classify_workbook,
)
from app.importers.workbook_profiler.workbook_scanner import scan_workbook
from app.schemas.import_schema import ColumnMappingSuggestion
from app.services.normalization_service import suggest_mapping
from app.utils.text_normalizer import normalize_text


MAX_SAMPLE_ROWS = 10
MAX_SAMPLE_COLUMNS = 12

FIELD_LABELS = {
    "driver_name": "Human Resource / driver",
    "second_driver_name": "Secondo driver",
    "vehicle_plate": "Targa / Asset",
    "station": "Operational Unit / station",
    "route": "Task / rotta",
    "cycle": "Time Window / wave",
    "status": "Stato",
    "workshop": "Officina",
    "notes": "Note",
    "key_available": "Chiave",
    "fuel_card": "Carta carburante",
    "vehicle_model": "Modello / categoria",
    "expirations": "Documenti / scadenze",
}

REQUIRED_FIELDS = {
    "planning": {"route", "station", "driver_name"},
    "fleet": {"vehicle_plate"},
}

COMPATIBLE_TYPES = {
    "planning": WorkbookType.DAILY_OPERATIONAL_PLANNING,
    "fleet": WorkbookType.FLEET_REGISTRY,
}


def _present(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def _json_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _columns_and_rows(
    sheet: ScannedSheet,
    header_row: int | None,
) -> tuple[list[str], list[dict[str, Any]], list[int], set[str]]:
    if not header_row or header_row > sheet.total_rows:
        return [], [], [], set()
    header = sheet.rows[header_row - 1]
    data_source = sheet.rows[header_row:]
    used_indexes = [
        index
        for index in range(sheet.total_columns)
        if (
            (index < len(header) and _present(header[index]))
            or any(index < len(row) and _present(row[index]) for row in data_source)
        )
    ]
    columns = []
    generic = set()
    seen: dict[str, int] = {}
    for index in used_indexes:
        raw = header[index] if index < len(header) else ""
        label = str(_json_value(raw)).strip()
        if not label:
            label = (
                f"Colonna Excel {get_column_letter(index + 1)} "
                "(intestazione assente)"
            )
            generic.add(label)
        key = normalize_text(label)
        seen[key] = seen.get(key, 0) + 1
        if seen[key] > 1:
            label = f"{label} ({get_column_letter(index + 1)})"
        columns.append(label)

    records = []
    row_numbers = []
    normalized_header = [normalize_text(item) for item in columns]
    for source_row, row in enumerate(data_source, start=header_row + 1):
        values = [
            row[index] if index < len(row) else ""
            for index in used_indexes
        ]
        present = sum(_present(value) for value in values)
        if present < min(2, max(1, len(columns))):
            continue
        if [normalize_text(value) for value in values] == normalized_header:
            continue
        records.append(
            {
                column: _json_value(value)
                for column, value in zip(columns, values)
            }
        )
        row_numbers.append(source_row)
    return columns, records, row_numbers, generic


def _apply_manual_mapping(
    suggestions: list[ColumnMappingSuggestion],
    manual_mapping: dict[str, str | None] | None,
    aliases: dict[str, list[str]],
    generic_columns: set[str],
) -> list[ColumnMappingSuggestion]:
    overrides = manual_mapping or {}
    unknown_sources = set(overrides) - {
        item.source_column for item in suggestions
    }
    if unknown_sources:
        raise WorkbookSelectionError(
            "Il mapping contiene colonne che non appartengono al foglio."
        )
    invalid_targets = {
        target
        for target in overrides.values()
        if target is not None and target not in aliases
    }
    if invalid_targets:
        raise WorkbookSelectionError(
            "Il mapping contiene campi incompatibili con il tipo di import."
        )
    selected_targets = [
        target for target in overrides.values() if target is not None
    ]
    if len(selected_targets) != len(set(selected_targets)):
        raise WorkbookSelectionError(
            "Ogni campo di destinazione puo essere usato una sola volta."
        )

    result = []
    for item in suggestions:
        source = item.source_column
        if source in overrides:
            target = overrides[source]
            status = "recognized" if target else "ignored"
            updated = item.model_copy(
                update={
                    "target_field": target,
                    "confidence": 1.0 if target else 0.0,
                    "requires_confirmation": False,
                    "status": status,
                }
            )
        elif source in generic_columns:
            updated = item.model_copy(
                update={
                    "target_field": None,
                    "confidence": 0.0,
                    "requires_confirmation": False,
                    "status": "ignored",
                }
            )
        elif item.target_field and not item.requires_confirmation:
            updated = item.model_copy(update={"status": "recognized"})
        elif item.target_field:
            updated = item.model_copy(update={"status": "review"})
        else:
            updated = item.model_copy(update={"status": "unknown"})

        result.append(updated)

    selected_by_target: dict[str, int] = {}
    for index, item in enumerate(result):
        if not item.target_field or item.status != "recognized":
            continue
        previous_index = selected_by_target.get(item.target_field)
        if previous_index is None:
            selected_by_target[item.target_field] = index
            continue
        previous = result[previous_index]
        item_rank = (
            item.source_column in overrides,
            item.confidence,
        )
        previous_rank = (
            previous.source_column in overrides,
            previous.confidence,
        )
        demoted_index = index
        if item_rank > previous_rank:
            selected_by_target[item.target_field] = index
            demoted_index = previous_index
        result[demoted_index] = result[demoted_index].model_copy(
            update={
                "target_field": None,
                "confidence": 0.0,
                "requires_confirmation": True,
                "status": "unknown",
            }
        )
    return result


def _issues(
    dataset_type: str,
    classification,
    selected_profile,
    selected_header,
    mapping,
    table_rows,
    sheet_count: int,
    column_count: int,
    generic_columns: set[str],
) -> tuple[list[ProfileIssue], list[ProfileIssue]]:
    blocking = []
    warnings = []
    expected = COMPATIBLE_TYPES[dataset_type]
    if classification.workbook_type != expected:
        if (
            dataset_type == "planning"
            and classification.workbook_type
            == WorkbookType.WORKFORCE_SCHEDULE
        ):
            message = (
                "Il file sembra contenere la programmazione dei turni "
                "driver, non il Planning operativo giornaliero con "
                "Task/rotte."
            )
        else:
            message = (
                f"Il workbook rilevato come "
                f"{classification.workbook_type.value} non e compatibile "
                f"con l'import {dataset_type}."
            )
        blocking.append(
            ProfileIssue(code="WORKBOOK_TYPE_MISMATCH", message=message)
        )
    if not selected_header:
        blocking.append(
            ProfileIssue(
                code="HEADER_NOT_FOUND",
                message="Non e stata individuata una riga intestazione.",
            )
        )
    elif (
        selected_header.confidence < 0.45
        and not selected_header.manually_selected
    ):
        blocking.append(
            ProfileIssue(
                code="HEADER_CONFIRMATION_REQUIRED",
                message=(
                    "La riga intestazione ha affidabilita bassa. "
                    "Selezionala manualmente e rianalizza."
                ),
            )
        )

    mapped_fields = {
        item.target_field
        for item in mapping
        if item.target_field and item.status == "recognized"
    }
    missing = REQUIRED_FIELDS[dataset_type] - mapped_fields
    if missing:
        labels = ", ".join(
            FIELD_LABELS.get(field, field)
            for field in sorted(missing)
        )
        blocking.append(
            ProfileIssue(
                code="REQUIRED_FIELDS_MISSING",
                message=f"Campi obbligatori mancanti: {labels}.",
            )
        )
    if not table_rows:
        blocking.append(
            ProfileIssue(
                code="NO_DATA_ROWS",
                message="Non sono state rilevate righe dati utilizzabili.",
            )
        )

    if selected_profile.formula_ratio:
        warnings.append(
            ProfileIssue(
                code="FORMULAS_PRESENT",
                message=(
                    "Il foglio contiene formule. Sono letti soltanto "
                    "i valori memorizzati dal workbook."
                ),
            )
        )
    if sheet_count > 10:
        warnings.append(
            ProfileIssue(
                code="MANY_SHEETS",
                message=f"Workbook complesso: {sheet_count} fogli rilevati.",
            )
        )
    if column_count > 40:
        warnings.append(
            ProfileIssue(
                code="WIDE_TABLE",
                message=(
                    f"Tabella molto larga: {column_count} colonne. "
                    "Il campione mostra soltanto quelle piu utili."
                ),
            )
        )
    if generic_columns:
        warnings.append(
            ProfileIssue(
                code="MISSING_HEADER_LABELS",
                message=(
                    f"{len(generic_columns)} colonne non hanno "
                    "un'intestazione esplicita."
                ),
            )
        )
    return blocking, warnings


def _sample_rows(
    rows: list[dict[str, Any]],
    row_numbers: list[int],
    mapping: list[ColumnMappingSuggestion],
) -> list[dict[str, Any]]:
    prioritized = [
        item.source_column
        for item in mapping
        if item.status in {"recognized", "review"}
    ]
    remaining = [
        item.source_column
        for item in mapping
        if item.source_column not in prioritized
        and item.status != "ignored"
    ]
    selected = (prioritized + remaining)[:MAX_SAMPLE_COLUMNS]
    return [
        {
            "Riga Excel": row_numbers[index],
            **{column: row.get(column, "") for column in selected},
        }
        for index, row in enumerate(rows[:MAX_SAMPLE_ROWS])
    ]


def build_workbook_profile(
    *,
    content: bytes,
    filename: str,
    dataset_type: str,
    aliases: dict[str, list[str]],
    sheet_name: str | None = None,
    header_row: int | None = None,
    manual_mapping: dict[str, str | None] | None = None,
) -> ProfiledWorkbook:
    if dataset_type not in REQUIRED_FIELDS:
        raise WorkbookSelectionError("Tipo dataset non supportato.")
    if header_row is not None and not 1 <= header_row <= 100:
        raise WorkbookSelectionError(
            "La riga intestazione deve essere compresa tra 1 e 100."
        )
    workbook = scan_workbook(content, filename)
    selected_sheet, selected_profile, profiles = profile_sheets(
        workbook,
        aliases,
        dataset_type,
        selected_sheet=sheet_name,
        manual_header_row=header_row,
    )
    selected_header = (
        selected_profile.header_candidates[0]
        if selected_profile.header_candidates
        else None
    )
    columns, table_rows, row_numbers, generic_columns = _columns_and_rows(
        selected_sheet,
        selected_header.row_index if selected_header else None,
    )
    mapping = _apply_manual_mapping(
        suggest_mapping(columns, aliases),
        manual_mapping,
        aliases,
        generic_columns,
    )
    classification = classify_workbook(
        workbook,
        selected_sheet,
        selected_header,
        confirmed_fields={
            item.target_field
            for item in mapping
            if item.target_field and item.status == "recognized"
        },
    )
    blocking, warnings = _issues(
        dataset_type,
        classification,
        selected_profile,
        selected_header,
        mapping,
        table_rows,
        len(workbook.sheets),
        len(columns),
        generic_columns,
    )
    recognized = [
        item for item in mapping if item.status == "recognized"
    ]
    ignored = [
        item.source_column for item in mapping if item.status == "ignored"
    ]
    unknown = [
        item.source_column
        for item in mapping
        if item.status in {"unknown", "review"}
    ]
    mapped_confidences = [
        item.confidence
        for item in mapping
        if item.target_field
    ]
    return ProfiledWorkbook(
        classification=classification,
        selected_sheet=selected_sheet,
        selected_sheet_profile=selected_profile,
        sheet_profiles=profiles,
        selected_header=selected_header,
        columns=columns,
        table_rows=table_rows,
        row_numbers=row_numbers,
        mapping=mapping,
        recognized_columns=recognized,
        ignored_columns=ignored,
        unknown_columns=unknown,
        mapping_options=[
            MappingFieldOption(
                value=field,
                label=FIELD_LABELS.get(field, field),
            )
            for field in aliases
        ],
        mapping_confidence=round(
            sum(mapped_confidences) / len(mapped_confidences),
            2,
        )
        if mapped_confidences
        else 0,
        import_allowed=not blocking,
        blocking_reasons=blocking,
        warnings=warnings,
        sample_rows=_sample_rows(table_rows, row_numbers, mapping),
    )
