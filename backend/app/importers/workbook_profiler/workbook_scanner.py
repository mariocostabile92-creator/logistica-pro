import csv
import io
import warnings
from datetime import datetime
from pathlib import Path
from time import perf_counter
from typing import Any

from openpyxl import load_workbook
import xlrd

from app.importers.workbook_profiler.errors import WorkbookReadError
from app.importers.workbook_profiler.models import (
    ScannedSheet,
    ScannedWorkbook,
)


def _is_present(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def _trim_matrix(
    rows: list[list[Any]],
    formula_cells: set[tuple[int, int]],
) -> list[list[Any]]:
    last_row = 0
    last_column = 0
    for row_index, row in enumerate(rows, start=1):
        for column_index, value in enumerate(row, start=1):
            if _is_present(value) or (row_index, column_index) in formula_cells:
                last_row = max(last_row, row_index)
                last_column = max(last_column, column_index)
    if not last_row or not last_column:
        return []
    return [
        list(row[:last_column]) + [""] * max(0, last_column - len(row))
        for row in rows[:last_row]
    ]


def _scan_csv(content: bytes) -> ScannedWorkbook:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t")
    except csv.Error:
        dialect = csv.excel
    rows = [list(row) for row in csv.reader(io.StringIO(text), dialect)]
    return ScannedWorkbook(
        sheets=(
            ScannedSheet(name="CSV", rows=_trim_matrix(rows, set())),
        )
    )


def _xls_value(book, cell) -> Any:
    if cell.ctype == xlrd.XL_CELL_DATE:
        return datetime(*xlrd.xldate_as_tuple(cell.value, book.datemode))
    return cell.value


def _scan_xls(content: bytes) -> ScannedWorkbook:
    book = xlrd.open_workbook(file_contents=content)
    sheets = []
    for source in book.sheets():
        rows = [
            [
                _xls_value(book, source.cell(row_index, column_index))
                for column_index in range(source.ncols)
            ]
            for row_index in range(source.nrows)
        ]
        merged = tuple(
            (row_low + 1, row_high, col_low + 1, col_high)
            for row_low, row_high, col_low, col_high
            in getattr(source, "merged_cells", ())
        )
        sheets.append(
            ScannedSheet(
                name=source.name,
                rows=_trim_matrix(rows, set()),
                merged_ranges=merged,
            )
        )
    return ScannedWorkbook(sheets=tuple(sheets))


def _formula_coordinates(worksheet) -> set[tuple[int, int]]:
    coordinates: set[tuple[int, int]] = set()
    for row in worksheet.iter_rows():
        for cell in row:
            value = cell.value
            if (
                cell.data_type == "f"
                or (isinstance(value, str) and value.startswith("="))
                or type(value).__name__.endswith("Formula")
            ):
                coordinates.add((cell.row, cell.column))
    return coordinates


def _scan_xlsx(
    content: bytes,
    *,
    preserve_formula_metadata: bool,
) -> ScannedWorkbook:
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Data Validation extension is not supported.*",
            category=UserWarning,
            module="openpyxl",
        )
        open_started = perf_counter()
        formulas_book = (
            load_workbook(
                io.BytesIO(content),
                read_only=False,
                data_only=False,
                keep_links=False,
            )
            if preserve_formula_metadata
            else None
        )
        values_book = load_workbook(
            io.BytesIO(content),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        open_seconds = perf_counter() - open_started
        try:
            scan_started = perf_counter()
            sheets = []
            for name in values_book.sheetnames:
                value_sheet = values_book[name]
                formula_sheet = formulas_book[name] if formulas_book else None
                formula_cells = (
                    _formula_coordinates(formula_sheet)
                    if formula_sheet is not None
                    else set()
                )
                max_row = max(
                    formula_sheet.max_row or 0
                    if formula_sheet is not None
                    else 0,
                    value_sheet.max_row or 0,
                )
                max_column = max(
                    formula_sheet.max_column or 0
                    if formula_sheet is not None
                    else 0,
                    value_sheet.max_column or 0,
                )
                rows = [
                    list(row)
                    for row in value_sheet.iter_rows(
                        min_row=1,
                        max_row=max_row,
                        min_col=1,
                        max_col=max_column,
                        values_only=True,
                    )
                ]
                merged = (
                    tuple(
                        (
                            item.min_row,
                            item.max_row,
                            item.min_col,
                            item.max_col,
                        )
                        for item in formula_sheet.merged_cells.ranges
                    )
                    if formula_sheet is not None
                    else ()
                )
                sheets.append(
                    ScannedSheet(
                        name=name,
                        rows=_trim_matrix(rows, formula_cells),
                        formula_cells=frozenset(formula_cells),
                        merged_ranges=merged,
                    )
                )
            return ScannedWorkbook(
                sheets=tuple(sheets),
                metrics={
                    "open_workbook": open_seconds,
                    "scan_sheets": perf_counter() - scan_started,
                },
            )
        finally:
            if formulas_book is not None:
                formulas_book.close()
            values_book.close()


def scan_workbook(
    content: bytes,
    filename: str,
    *,
    preserve_formula_metadata: bool = True,
) -> ScannedWorkbook:
    suffix = Path(filename).suffix.casefold()
    try:
        if suffix == ".csv":
            workbook = _scan_csv(content)
        elif suffix == ".xls":
            workbook = _scan_xls(content)
        else:
            workbook = _scan_xlsx(
                content,
                preserve_formula_metadata=preserve_formula_metadata,
            )
    except Exception as exc:
        raise WorkbookReadError(
            "Il file non e leggibile oppure la struttura Excel e danneggiata."
        ) from exc
    if not workbook.sheets:
        raise WorkbookReadError("Il workbook non contiene fogli.")
    return workbook
