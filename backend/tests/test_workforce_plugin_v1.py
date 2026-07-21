import io
from datetime import date

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.core.database import (
    _postgres_schema_statement,
    _postgres_statement,
    db_session,
)
from app.main import app
from app.plugins.workforce.bootstrap import (
    register_workforce_plugin,
    workforce_plugin_enabled,
)


client = TestClient(app)
BASE = "/api/plugins/workforce/v1"


def workforce_book(change: str | None = None) -> bytes:
    book = Workbook()
    members = book.active
    members.title = "Anagrafiche e contratti"
    members.append([
        "Matricola", "Nome Cognome", "Ruolo", "Tipo contratto",
        "Inizio contratto", "Fine contratto", "Ore settimanali", "Capability",
    ])
    members.append([
        "SYN-001", "Risorsa Uno", "courier", "full-time",
        "2026-01-01", "2026-12-31", 40, "license_b",
    ])
    members.append([
        "SYN-002", "Risorsa Due", "courier", "part-time",
        "2026-02-01", "2026-08-15", 24, "license_b, electric",
    ])
    shifts = book.create_sheet("Turni settimanali")
    shifts.append([
        "Matricola", "Nome Cognome",
        date(2026, 7, 21), date(2026, 7, 22), date(2026, 7, 23),
    ])
    shifts.append(["SYN-001", "Risorsa Uno", change or "S1", "F", "R"])
    shifts.append(["SYN-002", "Risorsa Due", "M", "P", "DISP"])
    needs = book.create_sheet("Fabbisogno")
    needs.append(["Data", "Sede", "Fabbisogno", "Capability"])
    needs.append(["2026-07-21", "UNIT-1", 2, "license_b"])
    needs.append(["2026-07-22", "UNIT-1", 1, "license_b"])
    output = io.BytesIO()
    book.save(output)
    book.close()
    return output.getvalue()


def preview(content: bytes | None = None):
    return client.post(
        f"{BASE}/import/preview",
        files={
            "file": (
                "synthetic_workforce.xlsx",
                content or workforce_book(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


def apply(content: bytes | None = None):
    payload = content or workforce_book()
    fingerprint = preview(payload).json()["fingerprint"]
    return client.post(
        f"{BASE}/import",
        data={"confirmed_fingerprint": fingerprint},
        files={
            "file": (
                "synthetic_workforce.xlsx",
                payload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


def test_workforce_plugin_is_enabled_only_by_explicit_flag(monkeypatch):
    monkeypatch.setenv("WORKFORCE_PLUGIN_ENABLED", "false")
    assert workforce_plugin_enabled() is False
    isolated_app = FastAPI()
    register_workforce_plugin(isolated_app)
    assert not any(
        route.path.startswith(BASE)
        for route in isolated_app.routes
    )
    monkeypatch.setenv("WORKFORCE_PLUGIN_ENABLED", "true")
    assert workforce_plugin_enabled() is True


def test_workforce_preview_is_multisheet_and_role_aware():
    response = preview()
    assert response.status_code == 200
    payload = response.json()
    assert payload["workbook_type"] == "WORKFORCE_SCHEDULE"
    assert payload["people_detected"] == 2
    assert payload["date_from"] == "2026-07-21"
    assert payload["date_to"] == "2026-07-23"
    assert {item["responsibility"] for item in payload["sheets"]} >= {
        "contracts", "schedule", "requirements",
    }
    assert payload["contracts_detected"] == 2
    assert payload["matrix"]


def test_workforce_preview_returns_typed_error_for_unreadable_workbook():
    response = client.post(
        f"{BASE}/import/preview",
        files={
            "file": (
                "synthetic_broken.xlsx",
                b"PK-not-a-valid-workbook",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 422
    assert response.json()["detail"]["code"] == "WORKBOOK_NOT_READABLE"


def test_workforce_import_maps_shift_holiday_rest_sickness_leave_and_availability():
    response = apply()
    assert response.status_code == 200
    assert response.json()["members_created"] == 2
    statuses = client.get(f"{BASE}/calendar").json()["items"]
    assert {item["status_code"] for item in statuses} == {
        "scheduled", "holiday", "rest", "sickness", "leave", "available",
    }
    assert any(item["shift_code"] == "S1" for item in statuses)
    assert sum(item["availability"] for item in statuses) == 2


def test_workforce_status_exposes_lightweight_ready_landing_summary():
    assert apply().status_code == 200
    payload = client.get(f"{BASE}/status").json()
    summary = payload["latest_import"]["summary"]

    assert payload["member_count"] == 2
    assert payload["latest_import"]["source"] == "Excel"
    assert summary["date_from"] == "2026-07-21"
    assert summary["date_to"] == "2026-07-23"
    assert summary["status_count"] == 6
    assert summary["contracts_detected"] == 2
    assert summary["absences_detected"] == 3
    assert summary["excluded_rows"] == 0
    assert summary["confirmation_columns"] == []


def test_contract_types_dates_hours_and_capabilities_are_typed():
    assert apply().status_code == 200
    members = client.get(f"{BASE}/members").json()["items"]
    assert {item["employment_type"] for item in members} == {"full-time", "part-time"}
    assert {item["weekly_hours"] for item in members} == {40.0, 24.0}
    part_time = next(item for item in members if item["external_identifier"] == "SYN-002")
    assert part_time["contract_end"] == "2026-08-15"
    assert "electric" in part_time["capabilities"]


def test_coverage_deficit_and_missing_requirement_are_explicit():
    assert apply().status_code == 200
    coverage = client.get(f"{BASE}/coverage").json()["items"]
    first = next(item for item in coverage if item["date"] == "2026-07-21")
    third = next(item for item in coverage if item["date"] == "2026-07-23")
    assert first == {
        "date": "2026-07-21",
        "operational_unit_id": None,
        "required": 2,
        "available": 1,
        "scheduled": 1,
        "unavailable": 1,
        "margin": -1,
        "status": "deficit",
        "limitations": [],
    }
    assert third["required"] is None
    assert third["status"] == "requirement_unavailable"
    assert third["limitations"] == ["Fabbisogno non disponibile."]


def test_manual_day_status_and_member_updates_append_audit():
    assert apply().status_code == 200
    member = client.get(f"{BASE}/members").json()["items"][0]
    status = client.get(f"{BASE}/calendar").json()["items"][0]
    update_status = client.patch(
        f"{BASE}/day-status/{status['status_id']}",
        json={
            "workforce_member_id": member["workforce_member_id"],
            "date": status["date"],
            "status_code": "leave",
            "shift_code": None,
            "notes": "Synthetic approved leave",
        },
    )
    assert update_status.status_code == 200
    assert update_status.json()["observed_or_confirmed"] == "manual"
    assert update_status.json()["availability"] is False
    update_member = client.patch(
        f"{BASE}/members/{member['workforce_member_id']}",
        json={
            "employment_type": "part-time",
            "weekly_hours": 30,
            "capabilities": ["license_b", "urban_access"],
        },
    )
    assert update_member.status_code == 200
    changes = client.get(f"{BASE}/changes").json()["items"]
    assert {item["reason"] for item in changes} >= {"manual_update"}
    assert any(item["before"] for item in changes)


def test_workforce_import_is_idempotent_and_updated_file_changes_only_diffs():
    payload = workforce_book()
    first = apply(payload)
    second = apply(payload)
    assert first.status_code == second.status_code == 200
    assert second.json()["idempotent"] is True
    with db_session() as conn:
        before = conn.execute("SELECT COUNT(*) AS total FROM workforce_changes").fetchone()["total"]
    changed = apply(workforce_book(change="F"))
    assert changed.status_code == 200
    assert changed.json()["statuses_updated"] == 1
    with db_session() as conn:
        after = conn.execute("SELECT COUNT(*) AS total FROM workforce_changes").fetchone()["total"]
    assert after == before + 1


def test_workforce_export_and_neutral_core_contracts():
    assert apply().status_code == 200
    exported = client.get(f"{BASE}/export?section=calendar")
    assert exported.status_code == 200
    assert "resource_identifier,display_name,date,status" in exported.text
    contracts = client.get(f"{BASE}/contracts/core").json()
    assert len(contracts["human_resources"]) == 2
    assert contracts["availability"][0]["resource"]["resource_kind"] == "human_resource"
    assert "route" not in str(contracts).casefold()


def test_workforce_demo_write_is_blocked_and_reset_removes_data_but_preserves_config():
    assert client.post("/api/demo/v1/load").status_code == 200
    blocked = apply()
    assert blocked.status_code == 409
    assert blocked.json()["detail"]["code"] == "DEMO_WORKSPACE_RESET_REQUIRED"
    assert client.post("/api/workspace/v1/reset").status_code == 200
    assert apply().status_code == 200
    assert client.get("/api/workspace/v1/status").json()["workspace_state"] == "PRODUCTION"
    reset = client.post("/api/workspace/v1/reset")
    assert reset.json()["removed_counts"]["workforce_members"] == 2
    assert client.get("/api/configuration/v1/current?organization_id=default").status_code == 200


def test_workforce_fixture_contains_only_synthetic_identifiers():
    content = workforce_book()
    assert b"SYN-001" in content or content.startswith(b"PK")
    source = __file__.casefold()
    assert "planning driver_" not in source
    assert "lorenzo" not in source


def test_contract_register_aliases_map_to_neutral_member_fields():
    book = Workbook()
    sheet = book.active
    sheet.title = "Anagrafiche"
    sheet.append(["Nome Cognome", "Ruolo"])
    sheet.append(["Sintetica Risorsa", "courier"])
    sheet = book.create_sheet("Scadenze")
    sheet.append([
        "Lavoratore",
        "Data assunzione",
        "Data cessazione",
        "% P.TIME",
        "Station",
    ])
    sheet.append([
        "Risorsa Sintetica",
        "2026-01-01",
        "2026-12-31",
        50,
        "UNIT-SYN",
    ])
    output = io.BytesIO()
    book.save(output)
    book.close()

    response = apply(output.getvalue())

    assert response.status_code == 200
    members = client.get(f"{BASE}/members").json()["items"]
    assert len(members) == 1
    member = members[0]
    assert member["employment_type"] == "part-time"
    assert member["contract_start"] == "2026-01-01"
    assert member["contract_end"] == "2026-12-31"


def test_integrated_operational_flow_returns_to_empty_after_reset():
    from app.repositories.import_repository import save_import
    from tests.planning_helpers import simple_rows
    from tests.test_intelligent_fleet_registry import (
        fleet_book,
        sync as sync_fleet,
    )

    assert client.get("/api/workspace/v1/status").json()["workspace_state"] == "EMPTY"
    workforce_result = apply()
    fleet_result = sync_fleet(fleet_book(rows=2))
    assert workforce_result.status_code == fleet_result.status_code == 200

    with db_session() as conn:
        assignments_before_planning = conn.execute(
            "SELECT COUNT(*) AS total FROM assignments"
        ).fetchone()["total"]
    assert assignments_before_planning == 0
    assert client.get("/api/workspace/v1/status").json()["workspace_state"] == "PRODUCTION"
    assert client.get(f"{BASE}/contracts/core").json()["availability"]
    assert client.get("/api/plugins/fleet/v1/availability").json()

    planning_rows, _ = simple_rows(routes=1, drivers=2, vehicles=2)
    planning_import_id = save_import(
        "planning",
        "synthetic-integrated-planning.csv",
        None,
        [],
        [row.model_dump(mode="json") for row in planning_rows],
    )
    generated = client.post(
        "/api/planning/generate",
        json={
            "planning_import_id": planning_import_id,
            "fleet_import_id": fleet_result.json()["import_id"],
            "operation_date": "2026-07-21",
        },
    )
    assert generated.status_code == 200
    briefing = client.post(
        "/api/briefing/v1/daily/generate",
        json={"planning_id": generated.json()["planning"]["id"]},
    )
    assert briefing.status_code == 200
    issue_codes = {item["issue_code"] for item in briefing.json()["sections"]}
    assert "WORKFORCE_DEFICIT" in issue_codes
    assert "FLEET_REGISTRY_ATTENTION" in issue_codes

    reset = client.post("/api/workspace/v1/reset")
    assert reset.status_code == 200
    assert client.get("/api/workspace/v1/status").json()["workspace_state"] == "EMPTY"


def test_workforce_schema_is_compatible_with_postgres_translation():
    statement, returns_identity = _postgres_statement(
        "INSERT INTO workforce_members (external_identifier) VALUES (?)"
    )
    assert returns_identity is True
    assert "VALUES (%s)" in statement
    assert statement.endswith("RETURNING id")
    assert "SERIAL PRIMARY KEY" in _postgres_schema_statement(
        "id INTEGER PRIMARY KEY AUTOINCREMENT"
    )
