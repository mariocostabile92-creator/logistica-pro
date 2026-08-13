import io
from datetime import date, timedelta

from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.core.database import db_session
from app.main import app
from app.plugins.workforce.application import coverage_service
from app.plugins.workforce.domain.coverage import (
    CoverageStatus,
    ImportedDailyCoverageRequirement,
)
from app.plugins.workforce.importer.workbook_interpreter import (
    interpret_workforce_workbook,
)
from app.plugins.workforce.infrastructure import coverage_repository
from app.utils.date_utils import utc_now_iso


BASE = "/api/plugins/workforce/v1"
client = TestClient(app)


def _planning_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Planning"
    sheet["H13"] = "FORECAST"
    sheet["H14"] = 0.1
    sheet["H15"] = "DRIVER DISPONIBILI C1"
    sheet["H16"] = "TESTE IN PIÙ O IN MENO DEL 10%"
    sheet["H19"] = "FORECAST SAME DAY A"
    sheet["H20"] = "FORECAST SAME DAY B - C"
    sheet["H21"] = "DRIVER DISPONIBILI SAME DAY A"
    sheet["H22"] = "DRIVER DISPONIBILI SAME DAY B-C"
    sheet["G24"] = "Turno"
    sheet["H24"] = "drivers"
    sheet["L24"] = date(2026, 8, 10)
    sheet["M24"] = date(2026, 8, 11)
    sheet["L13"] = 76
    sheet["M13"] = 0
    sheet["L14"] = "=L13*1.1"
    sheet["M14"] = "=M13*1.1"
    sheet["L19"] = 20
    sheet["L20"] = 18
    drivers = (
        ("NEXT", "Next Driver", "C1"),
        ("NEXT", "Second Next", "C1"),
        ("MATTINO", "Same A Driver", "SA"),
        ("POMERIGGIO", "Same BC Driver", "SB"),
    )
    for row, (group, name, shift) in enumerate(drivers, start=25):
        sheet.cell(row=row, column=7, value=group)
        sheet.cell(row=row, column=8, value=name)
        sheet.cell(row=row, column=12, value=shift)
        sheet.cell(row=row, column=13, value=shift)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _import(content: bytes):
    files = {
        "file": (
            "Planning driver_DLO2_2026 (1).xlsx",
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    preview = client.post(f"{BASE}/import/preview", files=files)
    assert preview.status_code == 200, preview.text
    files = {
        "file": (
            "Planning driver_DLO2_2026 (1).xlsx",
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }
    result = client.post(
        f"{BASE}/import",
        data={"confirmed_fingerprint": preview.json()["fingerprint"]},
        files=files,
    )
    assert result.status_code == 200, result.text
    return preview.json(), result.json()


def test_real_layout_parser_preserves_three_forecast_buckets_and_rounding():
    parsed = interpret_workforce_workbook(
        _planning_workbook(), "Planning driver_DLO2_2026 (1).xlsx"
    )
    indexed = {
        (
            item.operational_date,
            item.operational_cycle,
            item.coverage_segment,
        ): item
        for item in parsed.coverage_requirements
    }
    next_day = indexed[("2026-08-10", "NEXT_DAY", None)]
    assert next_day.forecast_routes == 76
    assert next_day.required_capacity == 84
    assert next_day.reserve_percentage == 10
    assert next_day.source_reference == "Planning!L13"
    assert indexed[("2026-08-10", "SAME_DAY", "A")].forecast_routes == 20
    assert indexed[("2026-08-10", "SAME_DAY", "B_C")].forecast_routes == 18
    assert indexed[("2026-08-11", "NEXT_DAY", None)].forecast_routes == 0
    assert parsed.preview.coverage_requirements_detected == 4


def test_real_layout_derives_member_cycle_from_dispatcher_group_without_guessing():
    parsed = interpret_workforce_workbook(
        _planning_workbook(), "planning.xlsx"
    )
    cycles = {item.values["display_name"]: item.values["operational_cycle"] for item in parsed.members}
    assert cycles["Next Driver"] == "NEXT_DAY"
    assert cycles["Same A Driver"] == "SAME_DAY"
    assert cycles["Same BC Driver"] == "SAME_DAY"


def test_requirement_math_covers_all_deterministic_statuses_and_zero_forecast():
    under = coverage_service._read_model(
        operational_date="2026-08-10", cycle="NEXT_DAY", segment=None,
        station=None, forecast_routes=76, reserve_percentage=10,
        required_capacity=84, assigned_drivers=70,
    )
    forecast_covered = coverage_service._read_model(
        operational_date="2026-08-10", cycle="NEXT_DAY", segment=None,
        station=None, forecast_routes=76, reserve_percentage=10,
        required_capacity=84, assigned_drivers=78,
    )
    requirement_covered = coverage_service._read_model(
        operational_date="2026-08-10", cycle="NEXT_DAY", segment=None,
        station=None, forecast_routes=76, reserve_percentage=10,
        required_capacity=84, assigned_drivers=84,
    )
    reserve = coverage_service._read_model(
        operational_date="2026-08-10", cycle="NEXT_DAY", segment=None,
        station=None, forecast_routes=76, reserve_percentage=10,
        required_capacity=84, assigned_drivers=86,
    )
    zero = coverage_service._read_model(
        operational_date="2026-08-11", cycle="NEXT_DAY", segment=None,
        station=None, forecast_routes=0, reserve_percentage=10,
        required_capacity=0, assigned_drivers=0,
    )
    assert (under.forecast_gap, under.requirement_gap, under.coverage_status) == (
        6, 14, CoverageStatus.UNDER_FORECAST
    )
    assert (forecast_covered.forecast_gap, forecast_covered.requirement_gap) == (0, 6)
    assert forecast_covered.coverage_status == CoverageStatus.FORECAST_COVERED
    assert requirement_covered.coverage_status == CoverageStatus.REQUIREMENT_COVERED
    assert reserve.reserve_drivers == 2
    assert zero.coverage_status == CoverageStatus.REQUIREMENT_COVERED


def test_import_persists_forecast_idempotently_and_endpoint_counts_assignments():
    content = _planning_workbook()
    preview, result = _import(content)
    assert preview["coverage_requirements_detected"] == 4
    assert result["coverage_requirements_created"] == 4
    response = client.get(
        f"{BASE}/planning/coverage",
        params={"date_from": "2026-08-10", "date_to": "2026-08-11"},
    )
    assert response.status_code == 200, response.text
    items = {
        (item["operational_date"], item["cycle"], item["segment"]): item
        for item in response.json()["items"]
    }
    assert items[("2026-08-10", "NEXT_DAY", None)]["assigned_drivers"] == 2
    assert items[("2026-08-10", "SAME_DAY", "A")]["assigned_drivers"] == 1
    assert items[("2026-08-10", "SAME_DAY", "B_C")]["assigned_drivers"] == 1
    assert items[("2026-08-11", "NEXT_DAY", None)]["forecast_routes"] == 0
    assert items[("2026-08-11", "SAME_DAY", "A")]["coverage_status"] == "NO_FORECAST"
    _, repeated = _import(content)
    assert repeated["idempotent"] is True
    with db_session() as conn:
        count = conn.execute(
            "SELECT COUNT(*) total FROM workforce_daily_coverage_requirements"
        ).fetchone()["total"]
    assert count == 4


def test_week_range_cycle_filter_and_no_forecast_are_explicit():
    response = client.get(
        f"{BASE}/planning/coverage",
        params={
            "date_from": "2026-08-10",
            "date_to": "2026-08-16",
            "cycle": "NEXT_DAY",
        },
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert len(body["items"]) == 7
    assert {item["cycle"] for item in body["items"]} == {"NEXT_DAY"}
    assert all(item["coverage_status"] == "NO_FORECAST" for item in body["items"])
    assert body["summary"]["no_forecast_buckets"] == 7


def test_organization_isolation_and_unknown_bucket_are_safe():
    now = utc_now_iso()
    requirement = ImportedDailyCoverageRequirement(
        operational_date="2026-08-10",
        station=None,
        operational_cycle="UNKNOWN_CYCLE",
        coverage_segment="LEGACY",
        forecast_routes=3,
        reserve_percentage=10,
        required_capacity=3,
        source="IMPORT",
        source_reference="Legacy!A1",
        source_identity="import:other",
    )
    with db_session() as conn:
        coverage_repository.persist_imported_requirements(
            conn, [requirement], organization_id="other-organization", now=now
        )
    response = client.get(
        f"{BASE}/planning/coverage",
        params={"date_from": "2026-08-10", "date_to": "2026-08-10"},
    )
    assert response.status_code == 200, response.text
    assert all(item["cycle"] != "UNKNOWN_CYCLE" for item in response.json()["items"])
    isolated = coverage_service.daily_coverage(
        "other-organization", "2026-08-10", "2026-08-10"
    )
    unknown = next(item for item in isolated.items if item.cycle == "UNKNOWN_CYCLE")
    assert unknown.forecast_routes == 3
    assert unknown.assigned_drivers == 0


def test_weekly_read_uses_one_requirement_query_and_one_assignment_query(monkeypatch):
    calls = {"requirements": 0, "assigned": 0}
    original_requirements = coverage_repository.list_current_requirements
    original_assigned = coverage_repository.assigned_driver_groups

    def requirements(*args, **kwargs):
        calls["requirements"] += 1
        return original_requirements(*args, **kwargs)

    def assigned(*args, **kwargs):
        calls["assigned"] += 1
        return original_assigned(*args, **kwargs)

    monkeypatch.setattr(coverage_repository, "list_current_requirements", requirements)
    monkeypatch.setattr(coverage_repository, "assigned_driver_groups", assigned)
    coverage_service.daily_coverage(
        "test-organization", "2026-08-10", "2026-08-16"
    )
    assert calls == {"requirements": 1, "assigned": 1}


def test_invalid_date_range_is_rejected():
    response = client.get(
        f"{BASE}/planning/coverage",
        params={"date_from": "2026-08-11", "date_to": "2026-08-10"},
    )
    assert response.status_code == 422


def test_all_week_dates_are_returned_in_chronological_order():
    start = date(2026, 8, 10)
    response = client.get(
        f"{BASE}/planning/coverage",
        params={
            "date_from": start.isoformat(),
            "date_to": (start + timedelta(days=6)).isoformat(),
        },
    )
    assert response.status_code == 200
    dates = [item["operational_date"] for item in response.json()["items"]]
    assert dates == sorted(dates)
