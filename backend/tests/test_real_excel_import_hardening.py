import hashlib
import io
import json
from datetime import datetime
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.adapters.amazon import AMAZON_ADAPTER
from app.core.database import _postgres_statement, db_session
from app.importers.workbook_profiler.models import WorkbookType
from app.importers.workbook_profiler.preview_builder import (
    build_workbook_profile,
)
from app.main import app
from app.repositories.import_repository import get_latest_import


MIME_XLSX = (
    "application/vnd.openxmlformats-officedocument."
    "spreadsheetml.sheet"
)
client = TestClient(app)


def workbook_bytes(build) -> bytes:
    workbook = Workbook()
    build(workbook)
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def simple_daily_workbook(*, header_row: int = 1) -> bytes:
    def build(workbook):
        sheet = workbook.active
        sheet.title = "Daily"
        for _ in range(header_row - 1):
            sheet.append([])
        sheet.append(["Route ID", "Delivery Station", "Driver", "Vehicle"])
        sheet.append(["TASK-001", "UNIT-01", "Resource 01", "AA001AA"])
        sheet.append(["TASK-002", "UNIT-01", "Resource 02", "AA002AA"])

    return workbook_bytes(build)


def fleet_workbook(*, header_row: int = 1) -> bytes:
    def build(workbook):
        sheet = workbook.active
        sheet.title = "Asset register"
        for _ in range(header_row - 1):
            sheet.append([])
        sheet.append(["Targa", "Modello", "Stato", "Officina", "Note"])
        sheet.append(["AA001AA", "Van A", "available", "", "Synthetic"])
        sheet.append(["AA002AA", "Van B", "maintenance", "Garage", "Synthetic"])

    return workbook_bytes(build)


def workforce_workbook() -> bytes:
    def build(workbook):
        sheet = workbook.active
        sheet.title = "Turni"
        sheet.append(
            [
                "Driver",
                "Contratto",
                "Lunedi",
                "Martedi",
                "Mercoledi",
                "Ferie",
            ]
        )
        sheet.append(["Resource 01", "Part time", "A", "A", "Riposo", ""])
        sheet.append(["Resource 02", "Full time", "B", "B", "B", ""])

    return workbook_bytes(build)


def profile(content: bytes, dataset_type: str, **overrides):
    return build_workbook_profile(
        content=content,
        filename="synthetic.xlsx",
        dataset_type=dataset_type,
        aliases=AMAZON_ADAPTER.aliases_for(dataset_type),
        **overrides,
    )


def post_preview(content: bytes, dataset_type: str, **data):
    return client.post(
        "/api/imports/preview",
        data={"dataset_type": dataset_type, **data},
        files={"file": ("synthetic.xlsx", content, MIME_XLSX)},
    )


def post_import(content: bytes, dataset_type: str, **data):
    return client.post(
        f"/api/imports/{dataset_type}",
        data=data,
        files={"file": ("synthetic.xlsx", content, MIME_XLSX)},
    )


def test_simple_daily_workbook_is_classified_and_explained():
    result = profile(simple_daily_workbook(), "planning")

    assert result.classification.workbook_type is (
        WorkbookType.DAILY_OPERATIONAL_PLANNING
    )
    assert result.classification.confidence >= 0.72
    assert "Task/rotte" in result.classification.reason
    assert result.import_allowed is True
    assert result.selected_header.row_index == 1


def test_multisheet_scoring_selects_tabular_sheet_and_reports_alternatives():
    def build(workbook):
        overview = workbook.active
        overview.title = "Overview"
        overview.merge_cells("A1:F2")
        overview["A1"] = "Synthetic operations report"
        for index in range(12):
            sheet = workbook.create_sheet(f"Decorative {index + 1}")
            sheet["A1"] = "Summary"
        daily = workbook.create_sheet("Daily operations")
        daily.append(["Route ID", "Delivery Station", "Driver", "Wave"])
        daily.append(["TASK-001", "UNIT-01", "Resource 01", "W1"])
        daily.append(["TASK-002", "UNIT-01", "Resource 02", "W2"])

    result = profile(workbook_bytes(build), "planning")

    assert result.selected_sheet.name == "Daily operations"
    assert result.selected_sheet_profile.score > 0.5
    assert len(result.sheet_profiles) == 14
    assert any(item.ignored for item in result.sheet_profiles)
    assert any(item.code == "MANY_SHEETS" for item in result.warnings)


def test_offset_header_merged_title_and_decorative_rows_are_ignored():
    def build(workbook):
        sheet = workbook.active
        sheet.title = "Fleet data"
        sheet.merge_cells("A1:E2")
        sheet["A1"] = "Synthetic fleet summary"
        sheet.append([])
        sheet.append(["Generated", datetime(2026, 7, 20)])
        sheet.append([])
        sheet.append(["Targa", "Modello", "Stato", "Officina", "Note"])
        sheet.append(["AA001AA", "Van A", "available", "", "Synthetic"])
        sheet.append(["AA002AA", "Van B", "reserve", "", "Synthetic"])

    result = profile(workbook_bytes(build), "fleet")

    assert result.selected_header.row_index == 5
    assert result.selected_sheet_profile.data_rows == 2
    assert result.table_rows[0]["Targa"] == "AA001AA"


def test_formula_cells_are_reported_without_becoming_random_headers():
    def build(workbook):
        sheet = workbook.active
        sheet.append(
            ["Route ID", "Delivery Station", "Driver", "Vehicle", "Check"]
        )
        sheet.append(["TASK-001", "UNIT-01", "Resource 01", "AA001AA", "=1+1"])
        sheet.append(["TASK-002", "UNIT-01", "Resource 02", "AA002AA", "=1+1"])

    result = profile(workbook_bytes(build), "planning")

    assert result.selected_header.row_index == 1
    assert result.import_allowed is True
    assert any(item.code == "FORMULAS_PRESENT" for item in result.warnings)


@pytest.mark.parametrize(
    ("content", "dataset_type", "expected"),
    [
        (
            simple_daily_workbook(),
            "planning",
            WorkbookType.DAILY_OPERATIONAL_PLANNING,
        ),
        (
            workforce_workbook(),
            "planning",
            WorkbookType.WORKFORCE_SCHEDULE,
        ),
        (
            fleet_workbook(),
            "fleet",
            WorkbookType.FLEET_REGISTRY,
        ),
    ],
)
def test_supported_workbook_types_are_deterministic(
    content,
    dataset_type,
    expected,
):
    first = profile(content, dataset_type).classification
    second = profile(content, dataset_type).classification

    assert first.workbook_type is expected
    assert first == second
    assert first.reason


def test_unknown_workbook_returns_analysis_instead_of_http_500():
    def build(workbook):
        sheet = workbook.active
        sheet.append(["Alpha", "Beta"])
        sheet.append(["one", "two"])

    response = post_preview(workbook_bytes(build), "planning")

    assert response.status_code == 200
    payload = response.json()
    assert payload["workbook_type"] == "UNKNOWN_WORKBOOK"
    assert payload["import_allowed"] is False
    assert payload["blocking_reasons"]


def test_manual_sheet_and_header_selection_are_honored():
    def build(workbook):
        workbook.active.title = "Read me"
        workbook.active["A1"] = "Synthetic instructions"
        sheet = workbook.create_sheet("Manual table")
        for _ in range(7):
            sheet.append([])
        sheet.append(["Route ID", "Delivery Station", "Driver"])
        sheet.append(["TASK-001", "UNIT-01", "Resource 01"])

    result = profile(
        workbook_bytes(build),
        "planning",
        sheet_name="Manual table",
        header_row=8,
    )

    assert result.selected_sheet.name == "Manual table"
    assert result.selected_header.row_index == 8
    assert result.selected_header.manually_selected is True
    assert result.import_allowed is True


def test_manual_mapping_accepts_only_compatible_targets():
    def build(workbook):
        sheet = workbook.active
        sheet.append(["Job Ref", "Depot Ref", "Operator Ref"])
        sheet.append(["TASK-001", "UNIT-01", "Resource 01"])

    content = workbook_bytes(build)
    mapping = {
        "Job Ref": "route",
        "Depot Ref": "station",
        "Operator Ref": "driver_name",
    }
    accepted = profile(content, "planning", manual_mapping=mapping)

    assert accepted.import_allowed is True
    with pytest.raises(ValueError, match="incompatibili"):
        profile(
            content,
            "planning",
            manual_mapping={"Job Ref": "document_type"},
        )


def test_workforce_schedule_is_previewed_but_blocked_from_planning_import():
    preview = post_preview(workforce_workbook(), "planning")
    imported = post_import(workforce_workbook(), "planning")

    assert preview.status_code == 200
    assert preview.json()["import_allowed"] is False
    assert any(
        item["code"] == "WORKBOOK_TYPE_MISMATCH"
        for item in preview.json()["blocking_reasons"]
    )
    assert imported.status_code == 422
    assert imported.json()["detail"]["code"] == "WORKBOOK_IMPORT_BLOCKED"
    assert get_latest_import("planning") is None


def test_valid_planning_and_fleet_imports_are_persisted_in_sqlite():
    planning = post_import(simple_daily_workbook(), "planning")
    fleet = post_import(fleet_workbook(), "fleet")

    assert planning.status_code == 200
    assert planning.json()["workbook_type"] == "DAILY_OPERATIONAL_PLANNING"
    assert fleet.status_code == 200
    assert fleet.json()["workbook_type"] == "FLEET_REGISTRY"
    assert get_latest_import("planning")["normalized_rows"]
    assert get_latest_import("fleet")["normalized_rows"]


def test_import_failure_rolls_back_without_partial_record(monkeypatch):
    from app.services import import_service

    def fail_save(**_kwargs):
        raise RuntimeError("synthetic persistence failure")

    monkeypatch.setattr(import_service, "save_import", fail_save)
    with pytest.raises(RuntimeError, match="synthetic persistence failure"):
        post_import(simple_daily_workbook(), "planning")
    with db_session() as conn:
        count = conn.execute(
            "SELECT COUNT(*) AS total FROM imports"
        ).fetchone()["total"]
    assert count == 0


def test_expected_empty_formula_and_corrupt_cases_never_return_http_500():
    def empty_build(_workbook):
        return None

    empty = post_preview(workbook_bytes(empty_build), "planning")
    corrupt = post_preview(b"not-an-excel-workbook", "planning")
    formula = post_preview(simple_daily_workbook(), "planning")

    assert empty.status_code == 200
    assert corrupt.status_code == 400
    assert formula.status_code == 200
    assert all(
        response.status_code != 500
        for response in (empty, corrupt, formula)
    )


def test_datetime_cell_regression_is_json_serializable():
    def build(workbook):
        sheet = workbook.active
        sheet.append(
            [
                "Route ID",
                "Delivery Station",
                "Driver",
                "Operation date",
            ]
        )
        sheet.append(
            [
                "TASK-001",
                "UNIT-01",
                "Resource 01",
                datetime(2026, 7, 20, 8, 30),
            ]
        )

    response = post_import(workbook_bytes(build), "planning")

    assert response.status_code == 200
    raw = response.json()["normalized_rows"][0]["raw"]
    assert raw["Operation date"] == "2026-07-20T08:30:00"


def test_adapter_and_configuration_aliases_feed_profiler():
    aliases = AMAZON_ADAPTER.aliases_for("planning")
    result = profile(simple_daily_workbook(), "planning")

    assert "route" in aliases
    assert "station" in aliases
    assert {
        item.target_field
        for item in result.mapping
        if item.target_field
    } >= {"route", "station", "driver_name"}


def test_postgresql_translation_keeps_import_insert_atomic_contract():
    statement, returns_identity = _postgres_statement(
        """
        INSERT INTO imports (
            dataset_type, original_filename, imported_at, sheet_name,
            column_mapping, normalized_rows
        ) VALUES (?, ?, ?, ?, ?, ?)
        """
    )

    assert statement.count("%s") == 6
    assert statement.rstrip().endswith("RETURNING id")
    assert returns_identity is True


def test_preexisting_non_import_openapi_paths_are_unchanged():
    paths = {
        path: value
        for path, value in app.openapi()["paths"].items()
        if not path.startswith("/api/imports/")
        and path != "/api/runtime/authority"
        and path != "/api/runtime/execution-intent"
        and not path.startswith("/api/plugins/workforce/")
        and not path.startswith("/api/planning/drafts")
        and not path.startswith("/api/planning/confirmation")
        and not path.startswith("/api/planning/publication")
        and path not in {
            "/api/planning/readiness",
            "/api/planning/conflicts",
            "/api/planning/timeline",
            "/api/plugins/fleet/v1/sync/preview",
            "/api/plugins/fleet/v1/sync/confirm",
            "/api/plugins/fleet/v1/sync/latest",
            "/api/plugins/fleet/v1/availability",
        }
    }
    digest = hashlib.sha256(
        json.dumps(
            paths,
            sort_keys=True,
            separators=(",", ":"),
        ).encode()
    ).hexdigest()

    assert digest == (
        "a17e205196c05b552a4aabd14f419cdc8dfc1e0b34bad3018edac9c85cde6455"
    )


def test_repository_contains_no_real_excel_regression_fixture():
    fixture_root = Path(__file__).with_name("fixtures")

    assert not list(fixture_root.rglob("*.xlsx"))
    assert not list(fixture_root.rglob("*.xls"))
