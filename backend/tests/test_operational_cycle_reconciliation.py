import io
import json
from hashlib import sha256
from uuid import uuid4

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.core.database import db_session
from app.main import app
from app.plugins.workforce.application import (
    operational_cycle_reconciliation_service as service,
)
from app.plugins.workforce.domain.operational_cycle_reconciliation import (
    OperationalCycleResolutionStatus,
)
from app.plugins.workforce.importer.workbook_interpreter import (
    interpret_workforce_workbook,
)
from app.plugins.workforce.infrastructure import (
    operational_cycle_reconciliation_repository as repository,
)


ORG = "test-organization"
BASE = "/api/plugins/workforce/v1/operational-cycle-backfill"
NOW = "2026-08-14T10:00:00Z"
client = TestClient(app)


def _workbook(rows: list[tuple[str, str, str | None]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Planning"
    sheet["D24"] = "T-ID"
    sheet["G24"] = "Turno"
    sheet["H24"] = "drivers"
    for index, (transporter, name, cycle) in enumerate(rows, start=25):
        sheet.cell(index, 4, transporter)
        sheet.cell(index, 7, cycle)
        sheet.cell(index, 8, name)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _source_rows(content: bytes):
    parsed = interpret_workforce_workbook(content, "planning.xlsx")
    return [
        row for row in parsed.source_rows
        if row.source_sheet == "Planning" and row.row_kind == "identity"
    ]


def _import(content: bytes, organization_id: str = ORG) -> int:
    with db_session() as conn:
        cursor = conn.execute(
            """
            INSERT INTO workforce_imports (
                fingerprint, original_filename, imported_at, sheets,
                summary, organization_id
            ) VALUES (?, 'Planning driver_DLO2_2026.xlsx', ?, '[]', '{}', ?)
            """,
            (sha256(content).hexdigest(), NOW, organization_id),
        )
        return int(cursor.lastrowid)


def _member(
    external_identifier: str,
    display_name: str,
    *,
    cycle: str = "NOT_SET",
    organization_id: str = ORG,
) -> int:
    with db_session() as conn:
        cursor = conn.execute(
            """
            INSERT INTO workforce_members (
                external_identifier, display_name, capabilities, active,
                source_reference, created_at, updated_at, organization_id,
                operational_cycle, station
            ) VALUES (?, ?, '[]', 1, 'legacy-planning', ?, ?, ?, ?, 'DLO2')
            """,
            (external_identifier, display_name, NOW, NOW, organization_id, cycle),
        )
        return int(cursor.lastrowid)


def _identity(transporter: str, member_id: int, organization_id: str = ORG):
    with db_session() as conn:
        conn.execute(
            """
            INSERT INTO workforce_external_identities (
                id, organization_id, source, external_id,
                workforce_member_id, status, created_at, updated_at
            ) VALUES (?, ?, 'amazon_transporter', ?, ?, 'MATCHED', ?, ?)
            """,
            (str(uuid4()), organization_id, transporter, member_id, NOW, NOW),
        )


def _prepared(
    rows: list[tuple[str, str, str | None]],
    *,
    cycles: dict[str, str] | None = None,
):
    content = _workbook(rows)
    import_id = _import(content)
    sources = _source_rows(content)
    members = {}
    for source in sources:
        if source.driver_display_name in members:
            continue
        members[source.driver_display_name] = _member(
            source.resolution_identifier,
            source.driver_display_name,
            cycle=(cycles or {}).get(source.driver_display_name, "NOT_SET"),
        )
    return content, import_id, sources, members


def _preview(content: bytes, import_id: int):
    return service.preview(
        ORG, content=content, filename="planning.xlsx",
        workforce_import_id=import_id,
    )


def test_exact_identity_resolution_uses_existing_workforce_identifier():
    content, import_id, sources, members = _prepared([
        ("T-1", "Mario Rossi", "NEXT"),
    ])
    result = _preview(content, import_id)
    detail = result.details[0]
    assert detail.status == OperationalCycleResolutionStatus.RESOLVED
    assert detail.workforce_member_id == members["Mario Rossi"]
    assert detail.workforce_external_identifier == sources[0].resolution_identifier


def test_transporter_mapping_is_reused_for_identity_resolution():
    content = _workbook([("T-2", "Driver Due", "NEXT")])
    import_id = _import(content)
    member_id = _member("canonical-other", "Canonical Driver")
    _identity("T-2", member_id)
    detail = _preview(content, import_id).details[0]
    assert detail.workforce_member_id == member_id
    assert "amazon_transporter" in (detail.resolution_source or "")


def test_next_day_requires_explicit_next_evidence():
    content, import_id, *_ = _prepared([("T-3", "Driver Next", "NEXT")])
    detail = _preview(content, import_id).details[0]
    assert detail.evidence_value == "NEXT"
    assert detail.proposed_cycle == "NEXT_DAY"


def test_mattino_is_explicit_same_day_evidence():
    content, import_id, *_ = _prepared([("T-4", "Driver Mattino", "MATTINO")])
    assert _preview(content, import_id).details[0].proposed_cycle == "SAME_DAY"


def test_pomeriggio_is_explicit_same_day_evidence():
    content, import_id, *_ = _prepared([("T-5", "Driver Pomeriggio", "POMERIGGIO")])
    assert _preview(content, import_id).details[0].proposed_cycle == "SAME_DAY"


@pytest.mark.parametrize("evidence", ["SAME DAY A", "SAME DAY B-C"])
def test_same_day_segments_remain_one_operational_cycle(evidence):
    content, import_id, *_ = _prepared([("T-6", f"Driver {evidence}", evidence)])
    assert _preview(content, import_id).details[0].proposed_cycle == "SAME_DAY"


def test_ambiguous_identity_is_never_selected_arbitrarily():
    content, import_id, sources, _ = _prepared([("T-7", "Driver Ambiguous", "NEXT")])
    other = _member("other-member", "Other Member")
    _identity("T-7", other)
    detail = _preview(content, import_id).details[0]
    assert detail.status == OperationalCycleResolutionStatus.AMBIGUOUS
    assert detail.apply_eligible is False


def test_not_found_does_not_create_a_workforce_member():
    content = _workbook([("T-8", "Missing Driver", "NEXT")])
    import_id = _import(content)
    before = _count("workforce_members")
    detail = _preview(content, import_id).details[0]
    assert detail.status == OperationalCycleResolutionStatus.NOT_FOUND
    assert _count("workforce_members") == before


def test_missing_explicit_cycle_is_no_cycle_evidence():
    content, import_id, *_ = _prepared([("T-9", "No Cycle", None)])
    detail = _preview(content, import_id).details[0]
    assert detail.status == OperationalCycleResolutionStatus.NO_CYCLE_EVIDENCE
    assert detail.proposed_cycle is None


def test_conflicting_explicit_cycles_for_one_member_are_rejected():
    content, import_id, *_ = _prepared([
        ("T-10A", "Same Driver", "NEXT"),
        ("T-10B", "Same Driver", "MATTINO"),
    ])
    details = _preview(content, import_id).details
    assert {item.status for item in details} == {
        OperationalCycleResolutionStatus.CONFLICT
    }
    assert not any(item.apply_eligible for item in details)


def test_existing_next_day_is_preserved_as_unchanged():
    content, import_id, *_ = _prepared(
        [("T-11", "Existing Next", "NEXT")],
        cycles={"Existing Next": "NEXT_DAY"},
    )
    result = _preview(content, import_id)
    assert result.summary.unchanged_existing_cycles == 1
    assert result.summary.apply_eligible == 0


def test_existing_same_day_is_preserved_as_unchanged():
    content, import_id, *_ = _prepared(
        [("T-12", "Existing Same", "POMERIGGIO")],
        cycles={"Existing Same": "SAME_DAY"},
    )
    assert _preview(content, import_id).summary.unchanged_existing_cycles == 1


def test_existing_cycle_disagreement_is_conflict():
    content, import_id, *_ = _prepared(
        [("T-13", "Existing Conflict", "NEXT")],
        cycles={"Existing Conflict": "SAME_DAY"},
    )
    assert _preview(content, import_id).details[0].status == (
        OperationalCycleResolutionStatus.CONFLICT
    )


def test_preview_fingerprint_is_stable_and_complete():
    content, import_id, *_ = _prepared([("T-14", "Fingerprint", "NEXT")])
    first = _preview(content, import_id)
    second = _preview(content, import_id)
    assert first.preview_fingerprint == second.preview_fingerprint
    assert len(first.preview_fingerprint or "") == 64


def test_stale_preview_fingerprint_is_rejected():
    content, import_id, *_ = _prepared([("T-15", "Stale", "NEXT")])
    with pytest.raises(service.OperationalCycleReconciliationConflictError):
        service.apply(
            ORG, content=content, filename="planning.xlsx",
            workforce_import_id=import_id,
            expected_preview_fingerprint="0" * 64, actor="admin@test.local",
        )


def test_apply_updates_only_resolved_not_set_members():
    content = _workbook([
        ("T-16A", "Eligible", "NEXT"),
        ("T-16B", "Missing", "MATTINO"),
    ])
    import_id = _import(content)
    source = _source_rows(content)[0]
    eligible = _member(source.resolution_identifier, "Eligible")
    preview = _preview(content, import_id)
    result = service.apply(
        ORG, content=content, filename="planning.xlsx",
        workforce_import_id=import_id,
        expected_preview_fingerprint=preview.preview_fingerprint or "",
        actor="admin@test.local",
    )
    assert result.members_updated == 1
    assert _cycle_for(eligible) == "NEXT_DAY"
    assert _count("workforce_members") == 1


def test_second_apply_with_current_preview_is_idempotent():
    content, import_id, _, members = _prepared([("T-17", "Idempotent", "NEXT")])
    first_preview = _preview(content, import_id)
    first = service.apply(
        ORG, content=content, filename="planning.xlsx",
        workforce_import_id=import_id,
        expected_preview_fingerprint=first_preview.preview_fingerprint or "",
        actor="admin@test.local",
    )
    second_preview = _preview(content, import_id)
    second = service.apply(
        ORG, content=content, filename="planning.xlsx",
        workforce_import_id=import_id,
        expected_preview_fingerprint=second_preview.preview_fingerprint or "",
        actor="admin@test.local",
    )
    assert first.members_updated == 1
    assert second.members_updated == 0
    assert second.idempotent is True
    assert _cycle_for(members["Idempotent"]) == "NEXT_DAY"


def test_organization_isolation_blocks_foreign_import_and_member():
    content = _workbook([("T-18", "Foreign Driver", "NEXT")])
    foreign_import = _import(content, "other-org")
    source = _source_rows(content)[0]
    _member(source.resolution_identifier, "Foreign Driver", organization_id="other-org")
    with pytest.raises(service.OperationalCycleReconciliationError):
        _preview(content, foreign_import)


def test_apply_is_atomic_when_second_update_fails(monkeypatch):
    content, import_id, _, members = _prepared([
        ("T-19A", "Atomic One", "NEXT"),
        ("T-19B", "Atomic Two", "MATTINO"),
    ])
    preview = _preview(content, import_id)
    original = repository._apply_member_cycle
    calls = 0

    def fail_second(*args, **kwargs):
        nonlocal calls
        calls += 1
        if calls == 2:
            raise RuntimeError("forced rollback")
        return original(*args, **kwargs)

    monkeypatch.setattr(repository, "_apply_member_cycle", fail_second)
    with pytest.raises(RuntimeError, match="forced rollback"):
        service.apply(
            ORG, content=content, filename="planning.xlsx",
            workforce_import_id=import_id,
            expected_preview_fingerprint=preview.preview_fingerprint or "",
            actor="admin@test.local",
        )
    assert {_cycle_for(member_id) for member_id in members.values()} == {"NOT_SET"}
    assert _count("workforce_changes") == 0


def test_audit_contains_bridge_provenance_without_raw_workbook():
    content, import_id, _, members = _prepared([("T-20", "Audited", "NEXT")])
    preview = _preview(content, import_id)
    result = service.apply(
        ORG, content=content, filename="planning.xlsx",
        workforce_import_id=import_id,
        expected_preview_fingerprint=preview.preview_fingerprint or "",
        actor="admin@test.local",
    )
    with db_session() as conn:
        row = conn.execute(
            "SELECT * FROM workforce_changes WHERE organization_id = ?",
            (ORG,),
        ).fetchone()
    after = json.loads(row["after_value"])
    assert result.audit_events_created == 1
    assert row["source"] == "LEGACY_CYCLE_RECONCILIATION"
    assert row["reason"] == "operational_cycle_changed"
    assert after["workforce_import_id"] == import_id
    assert after["source_reference"] == "Planning:row:25"
    assert "raw" not in after


def test_canonical_day_statuses_and_other_tables_are_unchanged():
    content, import_id, _, members = _prepared([("T-21", "Safe", "NEXT")])
    member_id = members["Safe"]
    _status(member_id, "2026-08-10", "C1")
    before = {
        name: _count(name) for name in (
            "workforce_day_statuses",
            "workforce_daily_coverage_requirements",
            "driver_shift_planning_published_rows",
            "driver_shift_distributions",
        )
    }
    preview = _preview(content, import_id)
    service.apply(
        ORG, content=content, filename="planning.xlsx",
        workforce_import_id=import_id,
        expected_preview_fingerprint=preview.preview_fingerprint or "",
        actor="admin@test.local",
    )
    assert {
        name: _count(name) for name in before
    } == before


def test_coverage_impact_uses_existing_bucket_semantics_without_writing():
    content, import_id, _, members = _prepared([
        ("T-22A", "Next Coverage", "NEXT"),
        ("T-22B", "Same A Coverage", "MATTINO"),
        ("T-22C", "Same B Coverage", "POMERIGGIO"),
    ])
    _status(members["Next Coverage"], "2026-08-10", "C1")
    _status(members["Same A Coverage"], "2026-08-10", "SA")
    _status(members["Same B Coverage"], "2026-08-10", "SB")
    _requirements("2026-08-10")
    result = _preview(content, import_id)
    indexed = {(item.cycle, item.segment): item for item in result.coverage_impact}
    assert indexed[("NEXT_DAY", None)].assigned_before == 0
    assert indexed[("NEXT_DAY", None)].assigned_after == 1
    assert indexed[("SAME_DAY", "A")].assigned_after == 1
    assert indexed[("SAME_DAY", "B_C")].assigned_after == 1
    assert {_cycle_for(value) for value in members.values()} == {"NOT_SET"}


def test_preview_endpoint_and_apply_endpoint_use_canonical_contract():
    content, import_id, *_ = _prepared([("T-23", "API Driver", "NEXT")])
    files = {"file": ("planning.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    preview = client.post(
        f"{BASE}/preview",
        data={"workforce_import_id": str(import_id)}, files=files,
    )
    assert preview.status_code == 200, preview.text
    applied = client.post(
        BASE,
        data={
            "workforce_import_id": str(import_id),
            "expected_preview_fingerprint": preview.json()["preview_fingerprint"],
        },
        files={"file": ("planning.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert applied.status_code == 200, applied.text
    assert applied.json()["members_updated"] == 1


def test_source_fingerprint_mismatch_is_rejected_before_resolution():
    content, import_id, *_ = _prepared([("T-24", "Source A", "NEXT")])
    changed = _workbook([("T-24", "Source A", "MATTINO")])
    with pytest.raises(service.OperationalCycleReconciliationError, match="fingerprint"):
        _preview(changed, import_id)


def _count(table: str) -> int:
    with db_session() as conn:
        return int(conn.execute(f"SELECT COUNT(*) total FROM {table}").fetchone()["total"])


def _cycle_for(member_id: int) -> str:
    with db_session() as conn:
        return str(conn.execute(
            "SELECT operational_cycle FROM workforce_members WHERE id = ?",
            (member_id,),
        ).fetchone()["operational_cycle"])


def _status(member_id: int, operational_date: str, shift_code: str):
    with db_session() as conn:
        conn.execute(
            """
            INSERT INTO workforce_day_statuses (
                workforce_member_id, date, status_code, availability,
                shift_code, source_reference, observed_or_confirmed,
                updated_at, organization_id
            ) VALUES (?, ?, 'scheduled', 1, ?, 'test', 'confirmed', ?, ?)
            """,
            (member_id, operational_date, shift_code, NOW, ORG),
        )


def _requirements(operational_date: str):
    with db_session() as conn:
        for cycle, segment in (
            ("NEXT_DAY", ""), ("SAME_DAY", "A"), ("SAME_DAY", "B_C")
        ):
            conn.execute(
                """
                INSERT INTO workforce_daily_coverage_requirements (
                    organization_id, operational_date, station, station_key,
                    operational_cycle, coverage_segment, forecast_routes,
                    reserve_percentage, required_capacity, source,
                    source_reference, source_identity, created_at, updated_at
                ) VALUES (?, ?, NULL, '', ?, ?, 10, 10, 11, 'TEST',
                          'fixture', ?, ?, ?)
                """,
                (
                    ORG, operational_date, cycle, segment,
                    f"fixture:{cycle}:{segment}", NOW, NOW,
                ),
            )
