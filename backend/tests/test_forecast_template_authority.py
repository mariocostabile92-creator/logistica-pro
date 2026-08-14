import io
import json
from collections import Counter
from dataclasses import replace
from datetime import date, timedelta
from hashlib import sha256
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.api.planning_workforce_bridge import planning_workforce_input
from app.auth.tenant_context import bind_organization, reset_organization
from app.core.database import db_session
from app.plugins.dsp_workspace.application.workforce_read_bridge import (
    coverage_projection,
)
from app.plugins.workforce.application import (
    coverage_service,
    forecast_reconciliation_service,
    legacy_coverage_backfill_service,
    workforce_service,
)
from app.plugins.workforce.domain.coverage import (
    CoverageSource,
    ForecastAuthorityStatus,
    ForecastDetectionReason,
    ImportedDailyCoverageRequirement,
    required_capacity_for,
)
from app.plugins.workforce.domain.forecast_authority import (
    MIN_TEMPLATE_RUN_DAYS,
    classify_forecast_requirements,
)
from app.plugins.workforce.importer.workbook_interpreter import (
    interpret_workforce_workbook,
)
from app.plugins.workforce.infrastructure import coverage_repository
from app.utils.date_utils import utc_now_iso


ORG = "test-organization"


def _requirement(
    offset: int,
    value: int,
    *,
    cycle: str = "NEXT_DAY",
    segment: str | None = None,
    source: str = CoverageSource.IMPORT.value,
    identity: str = "import:test",
) -> ImportedDailyCoverageRequirement:
    operational_date = (date(2026, 1, 1) + timedelta(days=offset)).isoformat()
    return ImportedDailyCoverageRequirement(
        operational_date=operational_date,
        station="DLO2",
        operational_cycle=cycle,
        coverage_segment=segment,
        forecast_routes=value,
        reserve_percentage=10,
        required_capacity=required_capacity_for(value),
        source=source,
        source_reference=f"Planning!{offset}",
        source_identity=identity,
    )


def _sequence(days: int, *, start: int = 50) -> list[ImportedDailyCoverageRequirement]:
    result = [_requirement(index, start + index) for index in range(days)]
    result.extend(
        _requirement(index, 20, cycle="SAME_DAY", segment="A")
        for index in range(days)
    )
    result.extend(
        _requirement(index, 18, cycle="SAME_DAY", segment="B_C")
        for index in range(days)
    )
    return result


def _workbook(days: int = 14, *, progressive: bool = True) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Planning"
    sheet["H13"] = "FORECAST"
    sheet["H14"] = 0.1
    sheet["H19"] = "FORECAST SAME DAY A"
    sheet["H20"] = "FORECAST SAME DAY B - C"
    sheet["G24"] = "Turno"
    sheet["H24"] = "drivers"
    start = date(2025, 12, 28)
    for offset in range(days):
        column = 12 + offset
        sheet.cell(row=24, column=column, value=start + timedelta(days=offset))
        sheet.cell(row=13, column=column, value=(59 + offset if progressive else 70))
        sheet.cell(row=19, column=column, value=20)
        sheet.cell(row=20, column=column, value=18)
    members = workbook.create_sheet("Anagrafiche")
    members.append(["Matricola", "Nome Cognome", "Station"])
    members.append(["WF-001", "Driver Test", "DLO2"])
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _legacy_import(content: bytes, *, organization_id: str = ORG) -> int:
    with db_session() as conn:
        cursor = conn.execute(
            """
            INSERT INTO workforce_imports (
                fingerprint, original_filename, imported_at, sheets,
                summary, organization_id
            ) VALUES (?, ?, ?, '[]', '{}', ?)
            """,
            (
                sha256(content).hexdigest(),
                "Planning driver_DLO2_2026.xlsx",
                "2026-08-08T10:23:10Z",
                organization_id,
            ),
        )
        return int(cursor.lastrowid)


def _persist(requirements, *, organization_id: str = ORG):
    with db_session() as conn:
        return coverage_repository.persist_imported_requirements(
            conn,
            requirements,
            organization_id=organization_id,
            now=utc_now_iso(),
        )


def _simulate_pre_authority_legacy_rows(
    requirements,
    *,
    organization_id: str = ORG,
) -> None:
    from app.plugins.workforce.infrastructure import legacy_coverage_backfill_repository
    legacy_coverage_backfill_repository.apply_missing(
        organization_id, requirements
    )
    with db_session() as conn:
        conn.execute(
            """
            UPDATE workforce_daily_coverage_requirements
            SET authority_status = 'AUTHORITATIVE', detection_reason = NULL
            WHERE organization_id = ? AND source = 'LEGACY_IMPORT_BACKFILL'
            """,
            (organization_id,),
        )


def _real_workbook() -> Path | None:
    candidates = (
        Path.home() / "Downloads" / "Planning driver_DLO2_2026 (1).xlsx",
        Path.home() / "Downloads" / "Planning driver_DLO2_2026(1).xlsx",
    )
    return next((path for path in candidates if path.exists()), None)


def test_detects_long_plus_one_sequence_and_correlated_constant_blocks():
    classified = classify_forecast_requirements(_sequence(20))
    counts = Counter(item.authority_status for item in classified)
    assert counts[ForecastAuthorityStatus.REJECTED_TEMPLATE.value] == 20
    assert counts[ForecastAuthorityStatus.SUSPECT_TEMPLATE.value] == 40
    assert all(
        item.detection_reason == ForecastDetectionReason.LONG_ARITHMETIC_SEQUENCE.value
        for item in classified
        if item.operational_cycle == "NEXT_DAY"
    )


def test_short_sequence_is_authoritative_and_exact_threshold_is_rejected():
    short = classify_forecast_requirements(_sequence(MIN_TEMPLATE_RUN_DAYS - 1))
    exact = classify_forecast_requirements(_sequence(MIN_TEMPLATE_RUN_DAYS))
    assert all(
        item.authority_status == ForecastAuthorityStatus.AUTHORITATIVE.value
        for item in short
    )
    assert sum(
        item.authority_status == ForecastAuthorityStatus.REJECTED_TEMPLATE.value
        for item in exact
    ) == MIN_TEMPLATE_RUN_DAYS


def test_non_pattern_zero_and_missing_dates_remain_valid():
    valid = [_requirement(0, 0), _requirement(1, 70), _requirement(3, 71)]
    classified = classify_forecast_requirements(valid)
    assert [item.forecast_routes for item in classified] == [0, 70, 71]
    assert all(
        item.authority_status == ForecastAuthorityStatus.AUTHORITATIVE.value
        for item in classified
    )


def test_import_preview_exposes_raw_authority_reason_and_effective_value():
    parsed = interpret_workforce_workbook(_workbook(), "planning.xlsx")
    assert parsed.preview.coverage_rejected_template == 14
    assert parsed.preview.coverage_suspect_template == 28
    rejected = next(
        item for item in parsed.preview.coverage_preview
        if item.cycle == "NEXT_DAY"
    )
    assert rejected.raw_forecast == 59
    assert rejected.authority_status == ForecastAuthorityStatus.REJECTED_TEMPLATE
    assert rejected.detection_reason == "LONG_ARITHMETIC_SEQUENCE"
    assert rejected.effective_forecast is None
    assert len(parsed.coverage_requirements) == 42
    assert any("template sono stati scartati" in item for item in parsed.preview.anomalies)


def test_confirmed_import_persists_partial_data_and_template_audit_without_raw_workbook():
    content = _workbook()
    token = bind_organization(ORG)
    try:
        preview = workforce_service.preview_import(content, "planning.xlsx")
        result = workforce_service.apply_import(
            content,
            "planning.xlsx",
            preview.fingerprint,
            actor="admin@example.test",
        )
    finally:
        reset_organization(token)
    assert result.members_created == 1
    assert result.coverage_requirements_created == 42
    with db_session() as conn:
        rows = conn.execute(
            """
            SELECT reason, after_value FROM workforce_changes
            WHERE organization_id = ?
              AND reason IN (
                'forecast_template_detected',
                'forecast_template_rejected'
              )
            ORDER BY reason
            """,
            (ORG,),
        ).fetchall()
    assert {row["reason"] for row in rows} == {
        "forecast_template_detected", "forecast_template_rejected"
    }
    assert all("raw" not in row["after_value"] for row in rows)


def test_rejected_is_non_effective_suspect_remains_effective_and_manual_wins():
    classified = classify_forecast_requirements(_sequence(14))
    _persist(classified)
    rejected_day = classified[0].operational_date
    initial = coverage_service.daily_coverage(ORG, rejected_day, rejected_day)
    next_day = next(item for item in initial.items if item.cycle == "NEXT_DAY")
    same_day = next(
        item for item in initial.items
        if item.cycle == "SAME_DAY" and item.segment == "A"
    )
    assert next_day.forecast_routes is None
    assert next_day.raw_forecast_routes == 50
    assert next_day.coverage_status.value == "NO_FORECAST"
    assert same_day.forecast_routes == 20
    assert same_day.authority_status == ForecastAuthorityStatus.SUSPECT_TEMPLATE

    manual = replace(
        classified[0],
        forecast_routes=70,
        required_capacity=77,
        source=CoverageSource.MANUAL_PLANNING_INPUT.value,
        source_identity="manual:2026-01-01:NEXT_DAY",
        authority_status=ForecastAuthorityStatus.AUTHORITATIVE.value,
        detection_reason=None,
    )
    _persist([manual])
    effective = coverage_service.daily_coverage(ORG, rejected_day, rejected_day)
    selected = next(item for item in effective.items if item.cycle == "NEXT_DAY")
    assert selected.forecast_routes == 70
    assert selected.source == CoverageSource.MANUAL_PLANNING_INPUT.value


def test_planning_workforce_and_dsp_share_the_same_non_effective_value():
    classified = classify_forecast_requirements(_sequence(14))
    _persist(classified)
    day = classified[0].operational_date
    workforce = coverage_service.daily_coverage(ORG, day, day)
    dsp_items, warnings = coverage_projection(workforce)
    planning = planning_workforce_input(operation_date=day, organization_id=ORG)
    workforce_next = next(item for item in workforce.items if item.cycle == "NEXT_DAY")
    dsp_next = next(item for item in dsp_items if item.cycle == "NEXT_DAY")
    planning_next = next(
        item for item in planning["coverage"]["items"] if item["cycle"] == "NEXT_DAY"
    )
    assert workforce_next.forecast_routes is None
    assert dsp_next.forecast is None
    assert planning_next["forecast"] is None
    assert any(item.code == "FORECAST_TEMPLATE_REJECTED" for item in warnings)


def test_reconciliation_is_organization_scoped_and_audited():
    content = _workbook()
    import_id = _legacy_import(content)
    inspection, requirements = legacy_coverage_backfill_service.inspect(
        ORG,
        content=content,
        filename="planning.xlsx",
        workforce_import_id=import_id,
    )
    assert inspection.status.value == "READY"
    _simulate_pre_authority_legacy_rows(requirements)

    result = forecast_reconciliation_service.preview(
        ORG,
        actor="admin@example.test",
        content=content,
        filename="planning.xlsx",
        workforce_import_id=import_id,
    )
    assert result.status.value == "READY"
    assert result.next_day_affected == 14
    assert result.same_day_a_suspect == 14
    assert result.same_day_b_c_suspect == 14
    assert result.effective_rows_before == 42
    assert result.effective_rows_after == 28
    assert result.rows_pending == 14
    foreign = forecast_reconciliation_service.preview(
        "other-organization",
        actor="admin@example.test",
        content=content,
        filename="planning.xlsx",
        workforce_import_id=import_id,
    )
    assert foreign.status.value == "NO_ELIGIBLE_IMPORT"
    with db_session() as conn:
        audit = conn.execute(
            """
            SELECT after_value FROM workforce_changes
            WHERE organization_id = ?
              AND reason = 'forecast_template_reconciliation_preview'
            """,
            (ORG,),
        ).fetchone()
    assert audit is not None
    assert "workbook" not in json.loads(audit["after_value"])


def test_reconciliation_apply_is_idempotent_preserves_manual_and_same_day():
    content = _workbook()
    import_id = _legacy_import(content)
    _, requirements = legacy_coverage_backfill_service.inspect(
        ORG,
        content=content,
        filename="planning.xlsx",
        workforce_import_id=import_id,
    )
    _simulate_pre_authority_legacy_rows(requirements)
    manual_source = next(
        item for item in requirements
        if item.operational_cycle == "NEXT_DAY"
    )
    _persist([
        replace(
            manual_source,
            forecast_routes=70,
            required_capacity=77,
            source=CoverageSource.MANUAL_PLANNING_INPUT.value,
            source_identity="manual:production-style:70",
            authority_status=ForecastAuthorityStatus.AUTHORITATIVE.value,
            detection_reason=None,
        )
    ])
    with db_session() as conn:
        manual_before = dict(conn.execute(
            """
            SELECT * FROM workforce_daily_coverage_requirements
            WHERE organization_id = ? AND source = 'MANUAL_PLANNING_INPUT'
            """,
            (ORG,),
        ).fetchone())
        same_day_before = [
            dict(row) for row in conn.execute(
                """
                SELECT * FROM workforce_daily_coverage_requirements
                WHERE organization_id = ? AND operational_cycle = 'SAME_DAY'
                ORDER BY id
                """,
                (ORG,),
            ).fetchall()
        ]

    preview = forecast_reconciliation_service.preview(
        ORG,
        actor="admin@example.test",
        content=content,
        filename="planning.xlsx",
        workforce_import_id=import_id,
    )
    applied = forecast_reconciliation_service.apply(
        ORG,
        actor="admin@example.test",
        content=content,
        filename="planning.xlsx",
        workforce_import_id=import_id,
        expected_preview_fingerprint=preview.preview_fingerprint or "",
    )
    assert applied.status.value == "ALREADY_COMPLETE"
    assert applied.rows_updated == 14
    assert applied.rows_unchanged == 0
    assert applied.idempotent is False
    assert applied.rows_pending == 0

    with db_session() as conn:
        manual_after = dict(conn.execute(
            """
            SELECT * FROM workforce_daily_coverage_requirements
            WHERE organization_id = ? AND source = 'MANUAL_PLANNING_INPUT'
            """,
            (ORG,),
        ).fetchone())
        same_day_after = [
            dict(row) for row in conn.execute(
                """
                SELECT * FROM workforce_daily_coverage_requirements
                WHERE organization_id = ? AND operational_cycle = 'SAME_DAY'
                ORDER BY id
                """,
                (ORG,),
            ).fetchall()
        ]
        rejected_rows = conn.execute(
            """
            SELECT forecast_routes, authority_status, detection_reason
            FROM workforce_daily_coverage_requirements
            WHERE organization_id = ? AND operational_cycle = 'NEXT_DAY'
              AND source = 'LEGACY_IMPORT_BACKFILL'
            """,
            (ORG,),
        ).fetchall()
        audits = conn.execute(
            """
            SELECT actor, after_value FROM workforce_changes
            WHERE organization_id = ?
              AND reason = 'forecast_template_reconciled'
            """,
            (ORG,),
        ).fetchall()
    assert manual_after == manual_before
    assert same_day_after == same_day_before
    assert {row["authority_status"] for row in rejected_rows} == {
        "REJECTED_TEMPLATE"
    }
    assert {row["detection_reason"] for row in rejected_rows} == {
        "LONG_ARITHMETIC_SEQUENCE"
    }
    assert {row["forecast_routes"] for row in rejected_rows} == set(range(59, 73))
    assert len(audits) == 1
    audit = json.loads(audits[0]["after_value"])
    assert audits[0]["actor"] == "admin@example.test"
    assert audit["affected_count"] == 14
    assert audit["rejected_count"] == 14
    assert audit["manual_overrides_preserved"] == 1
    assert "workbook" not in audit

    repeated = forecast_reconciliation_service.apply(
        ORG,
        actor="admin@example.test",
        content=content,
        filename="planning.xlsx",
        workforce_import_id=import_id,
        expected_preview_fingerprint=preview.preview_fingerprint or "",
    )
    assert repeated.status.value == "ALREADY_COMPLETE"
    assert repeated.rows_updated == 0
    assert repeated.rows_unchanged == 14
    assert repeated.idempotent is True
    with db_session() as conn:
        audit_count = conn.execute(
            """
            SELECT COUNT(*) total FROM workforce_changes
            WHERE organization_id = ?
              AND reason = 'forecast_template_reconciled'
            """,
            (ORG,),
        ).fetchone()["total"]
    assert audit_count == 1

    manual_day = manual_source.operational_date
    rejected_day = next(
        item.operational_date for item in requirements
        if item.operational_cycle == "NEXT_DAY"
        and item.operational_date != manual_day
    )
    manual_coverage = coverage_service.daily_coverage(
        ORG, manual_day, manual_day
    )
    rejected_coverage = coverage_service.daily_coverage(
        ORG, rejected_day, rejected_day
    )
    assert next(
        item for item in manual_coverage.items if item.cycle == "NEXT_DAY"
    ).forecast_routes == 70
    rejected_next = next(
        item for item in rejected_coverage.items if item.cycle == "NEXT_DAY"
    )
    assert rejected_next.forecast_routes is None
    assert rejected_next.raw_forecast_routes == 60
    planning = planning_workforce_input(
        operation_date=rejected_day, organization_id=ORG
    )
    assert next(
        item for item in planning["coverage"]["items"]
        if item["cycle"] == "NEXT_DAY"
    )["forecast"] is None
    dsp, _ = coverage_projection(rejected_coverage)
    assert next(item for item in dsp if item.cycle == "NEXT_DAY").forecast is None


def test_reconciliation_rejects_stale_fingerprint_and_is_atomic(monkeypatch):
    content = _workbook()
    import_id = _legacy_import(content)
    _, requirements = legacy_coverage_backfill_service.inspect(
        ORG,
        content=content,
        filename="planning.xlsx",
        workforce_import_id=import_id,
    )
    _simulate_pre_authority_legacy_rows(requirements)
    preview = forecast_reconciliation_service.preview(
        ORG,
        actor="admin@example.test",
        content=content,
        filename="planning.xlsx",
        workforce_import_id=import_id,
    )
    with db_session() as conn:
        conn.execute(
            """
            UPDATE workforce_daily_coverage_requirements
            SET forecast_routes = forecast_routes + 100
            WHERE id = (
                SELECT MIN(id) FROM workforce_daily_coverage_requirements
                WHERE organization_id = ? AND operational_cycle = 'NEXT_DAY'
            )
            """,
            (ORG,),
        )
    with pytest.raises(
        forecast_reconciliation_service.ForecastReconciliationConflictError
    ):
        forecast_reconciliation_service.apply(
            ORG,
            actor="admin@example.test",
            content=content,
            filename="planning.xlsx",
            workforce_import_id=import_id,
            expected_preview_fingerprint=preview.preview_fingerprint or "",
        )
    with db_session() as conn:
        conn.execute(
            """
            UPDATE workforce_daily_coverage_requirements
            SET forecast_routes = forecast_routes - 100
            WHERE id = (
                SELECT MIN(id) FROM workforce_daily_coverage_requirements
                WHERE organization_id = ? AND operational_cycle = 'NEXT_DAY'
            )
            """,
            (ORG,),
        )

    from app.plugins.workforce.infrastructure import forecast_reconciliation_repository
    original = forecast_reconciliation_repository._update_rows

    def fail_after_first(conn, rows):
        original(conn, rows[:1])
        raise RuntimeError("forced rollback")

    monkeypatch.setattr(
        forecast_reconciliation_repository, "_update_rows", fail_after_first
    )
    with pytest.raises(RuntimeError, match="forced rollback"):
        forecast_reconciliation_service.apply(
            ORG,
            actor="admin@example.test",
            content=content,
            filename="planning.xlsx",
            workforce_import_id=import_id,
            expected_preview_fingerprint=preview.preview_fingerprint or "",
        )
    with db_session() as conn:
        states = conn.execute(
            """
            SELECT DISTINCT authority_status
            FROM workforce_daily_coverage_requirements
            WHERE organization_id = ? AND operational_cycle = 'NEXT_DAY'
              AND source = 'LEGACY_IMPORT_BACKFILL'
            """,
            (ORG,),
        ).fetchall()
        audit = conn.execute(
            """
            SELECT COUNT(*) total FROM workforce_changes
            WHERE organization_id = ?
              AND reason = 'forecast_template_reconciled'
            """,
            (ORG,),
        ).fetchone()["total"]
    assert {row["authority_status"] for row in states} == {"AUTHORITATIVE"}
    assert audit == 0


@pytest.mark.skipif(_real_workbook() is None, reason="Workbook reale non disponibile")
def test_real_workbook_marks_exact_323_323_323_and_raw_239():
    path = _real_workbook()
    assert path is not None
    parsed = interpret_workforce_workbook(path.read_bytes(), path.name)
    counts = Counter(
        (item.operational_cycle, item.coverage_segment, item.authority_status)
        for item in parsed.coverage_requirements
    )
    assert counts[("NEXT_DAY", None, "REJECTED_TEMPLATE")] == 323
    assert counts[("SAME_DAY", "A", "SUSPECT_TEMPLATE")] == 323
    assert counts[("SAME_DAY", "B_C", "SUSPECT_TEMPLATE")] == 323
    august_14 = next(
        item for item in parsed.coverage_requirements
        if item.operational_date == "2026-08-14"
        and item.operational_cycle == "NEXT_DAY"
    )
    assert august_14.forecast_routes == 239
    assert august_14.authority_status == "REJECTED_TEMPLATE"


@pytest.mark.skipif(_real_workbook() is None, reason="Workbook reale non disponibile")
def test_real_legacy_reconciliation_previews_323_and_preserves_manual_70():
    path = _real_workbook()
    assert path is not None
    content = path.read_bytes()
    import_id = _legacy_import(content)
    inspection, requirements = legacy_coverage_backfill_service.inspect(
        ORG,
        content=content,
        filename=path.name,
        workforce_import_id=import_id,
    )
    assert inspection.next_day_rejected_count == 323
    _simulate_pre_authority_legacy_rows(requirements)
    august_14 = next(
        item for item in requirements
        if item.operational_date == "2026-08-14"
        and item.operational_cycle == "NEXT_DAY"
    )
    _persist([
        replace(
            august_14,
            forecast_routes=70,
            required_capacity=77,
            source=CoverageSource.MANUAL_PLANNING_INPUT.value,
            source_identity="manual:2026-08-14:NEXT_DAY",
            authority_status=ForecastAuthorityStatus.AUTHORITATIVE.value,
            detection_reason=None,
        )
    ])
    preview = forecast_reconciliation_service.preview(
        ORG,
        actor="admin@example.test",
        content=content,
        filename=path.name,
        workforce_import_id=import_id,
    )
    assert preview.next_day_affected == 323
    assert preview.same_day_a_suspect == 323
    assert preview.same_day_b_c_suspect == 323
    assert preview.manual_overrides_preserved == 1
    assert preview.period_start == "2026-02-15"
    assert preview.period_end == "2027-01-03"
    assert preview.effective_rows_before == 969
    assert preview.effective_rows_after == 647
    effective = coverage_service.daily_coverage(ORG, "2026-08-14", "2026-08-14")
    next_day = next(item for item in effective.items if item.cycle == "NEXT_DAY")
    assert next_day.forecast_routes == 70
    result = forecast_reconciliation_service.apply(
        ORG,
        actor="admin@example.test",
        content=content,
        filename=path.name,
        workforce_import_id=import_id,
        expected_preview_fingerprint=preview.preview_fingerprint or "",
    )
    assert result.rows_updated == 323
    assert result.manual_overrides_preserved == 1
    with db_session() as conn:
        counts = {
            (row["operational_cycle"], row["authority_status"]): row["total"]
            for row in conn.execute(
                """
                SELECT operational_cycle, authority_status, COUNT(*) total
                FROM workforce_daily_coverage_requirements
                WHERE organization_id = ? AND source = 'LEGACY_IMPORT_BACKFILL'
                GROUP BY operational_cycle, authority_status
                """,
                (ORG,),
            ).fetchall()
        }
    assert counts[("NEXT_DAY", "REJECTED_TEMPLATE")] == 323
    assert counts[("SAME_DAY", "AUTHORITATIVE")] == 739
