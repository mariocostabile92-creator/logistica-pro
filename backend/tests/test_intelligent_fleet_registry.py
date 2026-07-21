import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.core.database import db_session
from app.core.database import _postgres_schema_statement, _postgres_statement
from app.main import app
from app.plugins.fleet.infrastructure import sync_repository


client = TestClient(app)
BASE = "/api/plugins/fleet/v1"


def fleet_book(rows: int = 3, changed: bool = False, duplicate: bool = False) -> bytes:
    book = Workbook()
    sheet = book.active
    sheet.title = "Stato parco"
    sheet.append(["Registro sintetico"])
    sheet.append([])
    sheet.append([])
    sheet.append([])
    sheet.append([
        "Asset ID", "Targa", "Modello", "Stato", "Officina",
        "Sostitutivo", "Driver", "Documento", "Scadenza", "PIN carta",
    ])
    for index in range(rows):
        state = "Indisponibile" if changed and index == 0 else (
            "Officina" if index == 1 else "Riserva" if index == 2 else "Disponibile"
        )
        plate = "SY000AA" if duplicate and index == 1 else f"SY{index:03d}AA"
        sheet.append([
            f"ASSET-SYN-{index:03d}", plate, "light_van", state,
            "Centro tecnico" if index == 1 else "",
            "SY099ZZ" if index == 2 else "",
            f"RESOURCE-SYN-{index:03d}", "insurance",
            "2026-12-31", f"SECRET-{1000 + index}",
        ])
    output = io.BytesIO()
    book.save(output)
    book.close()
    return output.getvalue()


def preview(content: bytes | None = None):
    return client.post(
        f"{BASE}/sync/preview",
        files={
            "file": (
                "synthetic_fleet.xlsx",
                content or fleet_book(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


def sync(content: bytes | None = None, selected: list[int] | None = None):
    payload = content or fleet_book()
    proposal = preview(payload).json()
    rows = selected if selected is not None else [
        item["row_id"] for item in proposal["items"] if item["selected_by_default"]
    ]
    return client.post(
        f"{BASE}/sync/confirm",
        data={
            "confirmed_fingerprint": proposal["fingerprint"],
            "selected_rows": __import__("json").dumps(rows),
        },
        files={
            "file": (
                "synthetic_fleet.xlsx",
                payload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


def test_fleet_registry_preview_profiles_header_rows_and_sensitive_fields():
    response = preview()
    assert response.status_code == 200
    payload = response.json()
    assert payload["workbook_type"] == "FLEET_REGISTRY"
    assert payload["profiled_sheets"] == 1
    assert payload["selected_sheet"] == "Stato parco"
    assert payload["selected_header_row"] == 5
    assert payload["source_rows"] == 3
    assert payload["summary"]["sensitive_fields_excluded"] == 3
    assert all(
        field["reason"].startswith("Campo sensibile rilevato")
        for item in payload["items"] for field in item["sensitive_fields"]
    )
    assert "SECRET-" not in response.text


def test_sync_creates_asset_registry_snapshot_events_documents_and_core_availability():
    response = sync()
    assert response.status_code == 200
    payload = response.json()
    assert payload["created_assets"] == 3
    assert payload["events_created"] >= 3
    assert payload["documents_created"] == 3
    assets = client.get(f"{BASE}/assets").json()["items"]
    assert len(assets) == 3
    assert {item["availability"] for item in assets} == {"available", "maintenance", "reserve"}
    assert all(item["documents"] for item in assets)
    availability = client.get(f"{BASE}/availability").json()
    assert len(availability) == 3
    assert {item["resource_kind"] for item in availability} == {"asset"}
    with db_session() as conn:
        imported = conn.execute("SELECT normalized_rows FROM imports WHERE id = ?", (payload["import_id"],)).fetchone()
    assert len(__import__("json").loads(imported["normalized_rows"])) == 3


def test_driver_observation_remains_unresolved_and_does_not_create_workforce_member():
    assert sync().status_code == 200
    with db_session() as conn:
        metadata = conn.execute("SELECT observed_assigned_human_resource FROM fleet_asset_metadata ORDER BY asset_id").fetchall()
        workforce = conn.execute("SELECT COUNT(*) AS total FROM workforce_members").fetchone()["total"]
    assert metadata[0]["observed_assigned_human_resource"] == "RESOURCE-SYN-000"
    assert workforce == 0


def test_reimport_is_idempotent_without_duplicate_assets_or_events():
    content = fleet_book()
    first = sync(content)
    second = sync(content)
    assert first.status_code == second.status_code == 200
    assert second.json()["idempotent"] is True
    with db_session() as conn:
        assert conn.execute("SELECT COUNT(*) AS total FROM fleet_assets").fetchone()["total"] == 3
        assert conn.execute("SELECT COUNT(*) AS total FROM fleet_sync_runs").fetchone()["total"] == 1
        events = conn.execute("SELECT COUNT(*) AS total FROM fleet_asset_events").fetchone()["total"]
    assert events == first.json()["events_created"]

    proposal = preview(content).json()
    assert proposal["summary"]["unchanged_assets"] == 3
    assert {item["action"] for item in proposal["items"]} == {"NO_CHANGE"}


def test_changed_single_row_updates_only_matching_asset_and_event():
    assert sync().status_code == 200
    changed = sync(fleet_book(changed=True))
    assert changed.status_code == 200
    assert changed.json()["created_assets"] == 0
    assert changed.json()["updated_assets"] == 1
    assert changed.json()["events_created"] == 1
    assets = client.get(f"{BASE}/assets").json()["items"]
    assert assets[0]["availability"] == "unavailable"


def test_duplicate_plate_is_not_selected_or_applied():
    proposal = preview(fleet_book(duplicate=True)).json()
    duplicate = next(item for item in proposal["items"] if item["action"] == "POSSIBLE_DUPLICATE")
    assert duplicate["selected_by_default"] is False
    invalid_selection = sync(fleet_book(duplicate=True), [duplicate["row_id"]])
    assert invalid_selection.status_code == 422
    assert invalid_selection.json()["detail"]["code"] == "FLEET_SYNC_SELECTION_INVALID"


def test_invalid_plate_and_row_are_never_selected():
    book = Workbook()
    sheet = book.active
    sheet.title = "Stato parco"
    sheet.append(["Asset ID", "Targa", "Stato"])
    sheet.append(["ASSET-SYN-INVALID", "X", "Disponibile"])
    output = io.BytesIO()
    book.save(output)
    book.close()

    proposal = preview(output.getvalue()).json()

    assert proposal["summary"]["invalid_rows"] == 1
    assert proposal["items"][0]["action"] == "INVALID_ROW"
    assert proposal["items"][0]["selected_by_default"] is False


def test_exact_identity_conflict_is_detected_without_fuzzy_plate_matching():
    first = sync(fleet_book(rows=1))
    assert first.status_code == 200
    book = Workbook()
    sheet = book.active
    sheet.title = "Stato parco"
    sheet.append(["Asset ID", "Targa", "Stato"])
    sheet.append(["ASSET-SYN-000", "DIFFERENT9", "Disponibile"])
    output = io.BytesIO()
    book.save(output)
    book.close()
    proposal = preview(output.getvalue()).json()
    assert proposal["items"][0]["action"] == "CONFLICT"
    assert proposal["items"][0]["selected_by_default"] is False


def test_observed_alternative_identifier_is_used_for_exact_conflict_detection():
    assert sync(fleet_book(rows=1)).status_code == 200

    def identity_book(external_identifier: str, plate: str) -> bytes:
        book = Workbook()
        sheet = book.active
        sheet.title = "Stato parco"
        sheet.append(["Asset ID", "Targa", "Stato"])
        sheet.append([external_identifier, plate, "Disponibile"])
        output = io.BytesIO()
        book.save(output)
        book.close()
        return output.getvalue()

    alias_content = identity_book("ALT-SYN-000", "SY000AA")
    alias_sync = sync(alias_content)
    assert alias_sync.status_code == 200
    assert alias_sync.json()["updated_assets"] == 1
    with db_session() as conn:
        alternatives = conn.execute(
            "SELECT alternative_identifiers FROM fleet_asset_metadata"
        ).fetchone()["alternative_identifiers"]
    assert "ALT-SYN-000" in alternatives

    conflict = preview(identity_book("ALT-SYN-000", "SY999ZZ")).json()
    assert conflict["items"][0]["action"] == "CONFLICT"
    assert conflict["items"][0]["selected_by_default"] is False


def test_sensitive_values_are_never_persisted_in_registry_snapshot_or_events():
    assert sync().status_code == 200
    with db_session() as conn:
        values = []
        for table, columns in (
            ("imports", "column_mapping || normalized_rows"),
            ("fleet_asset_events", "details"),
            ("fleet_asset_metadata", "source_reference || alternative_identifiers"),
        ):
            rows = conn.execute(f"SELECT {columns} AS value FROM {table}").fetchall()
            values.extend(str(row["value"]) for row in rows)
    assert "SECRET-" not in " ".join(values)
    assert '"fuel_card": null' in " ".join(values)


def test_replacement_reference_and_document_expiry_remain_observations():
    assert sync().status_code == 200
    with db_session() as conn:
        metadata = conn.execute(
            "SELECT replacement_asset_reference FROM fleet_asset_metadata "
            "WHERE replacement_asset_reference IS NOT NULL"
        ).fetchall()
        documents = conn.execute(
            "SELECT document_type, expires_on FROM fleet_asset_documents"
        ).fetchall()

    assert [row["replacement_asset_reference"] for row in metadata] == ["SY099ZZ"]
    assert len(documents) == 3
    assert {row["document_type"] for row in documents} == {"insurance"}
    assert {row["expires_on"] for row in documents} == {"2026-12-31"}


def test_atomic_sync_rolls_back_assets_snapshot_and_events(monkeypatch):
    original = sync_repository._apply_item
    calls = 0

    def fail_second(*args, **kwargs):
        nonlocal calls
        calls += 1
        if calls == 2:
            raise RuntimeError("Synthetic transaction failure")
        return original(*args, **kwargs)

    monkeypatch.setattr(sync_repository, "_apply_item", fail_second)
    with pytest.raises(RuntimeError, match="Synthetic transaction failure"):
        sync()
    with db_session() as conn:
        for table in ("fleet_assets", "fleet_asset_events", "fleet_sync_runs", "imports"):
            assert conn.execute(f"SELECT COUNT(*) AS total FROM {table}").fetchone()["total"] == 0


def test_demo_workspace_blocks_real_fleet_sync_with_typed_409():
    assert client.post("/api/demo/v1/load").status_code == 200
    response = sync()
    assert response.status_code == 409
    assert response.json()["detail"]["code"] == "DEMO_WORKSPACE_RESET_REQUIRED"


def test_86_structural_rows_sync_without_duplicates_and_reset_to_empty():
    response = sync(fleet_book(rows=86))
    assert response.status_code == 200
    assert response.json()["created_assets"] == 86
    status = client.get("/api/workspace/v1/status").json()
    assert status["workspace_state"] == "PRODUCTION"
    assert status["asset_count"] == 86
    reset = client.post("/api/workspace/v1/reset")
    assert reset.status_code == 200
    assert reset.json()["removed_counts"]["fleet_assets"] == 86
    assert client.get("/api/workspace/v1/status").json()["workspace_state"] == "EMPTY"


def test_repeated_vehicle_models_never_merge_distinct_plates():
    book = Workbook()
    sheet = book.active
    sheet.title = "Stato parco"
    sheet.append(["Veicolo", "Targa", "Danno", "Officina"])
    sheet.append(["synthetic-model", "SY100AA", 0, None])
    sheet.append(["synthetic-model", "SY101AA", 0, None])
    output = io.BytesIO()
    book.save(output)
    book.close()

    response = sync(output.getvalue())

    assert response.status_code == 200
    assert response.json()["created_assets"] == 2
    assets = client.get(f"{BASE}/assets").json()["items"]
    assert {asset["external_identifier"] for asset in assets} == {
        "SY100AA",
        "SY101AA",
    }


def test_fleet_sync_schema_is_compatible_with_postgres_translation():
    statement, returns_identity = _postgres_statement(
        "INSERT INTO fleet_sync_runs (application_key) VALUES (?)"
    )
    assert returns_identity is True
    assert "VALUES (%s)" in statement
    assert statement.endswith("RETURNING id")
    assert "SERIAL PRIMARY KEY" in _postgres_schema_statement(
        "id INTEGER PRIMARY KEY AUTOINCREMENT"
    )
