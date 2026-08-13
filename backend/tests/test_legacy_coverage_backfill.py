import io
import json
from datetime import date, timedelta
from hashlib import sha256

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from app.core.database import db_session
from app.main import app
from app.plugins.workforce.application import legacy_coverage_backfill_service
from app.plugins.workforce.domain.coverage import ImportedDailyCoverageRequirement
from app.plugins.workforce.infrastructure import (
    coverage_repository,
    legacy_coverage_backfill_repository,
)
from app.utils.date_utils import utc_now_iso


BASE = "/api/plugins/workforce/v1/planning/coverage/backfill"
ORG = "test-organization"
client = TestClient(app)


def _planning_workbook(
    *,
    days: int = 2,
    missing_next_day: set[int] | None = None,
    missing_same_day_a: set[int] | None = None,
    missing_same_day_b_c: set[int] | None = None,
) -> bytes:
    missing_next_day = missing_next_day or set()
    missing_same_day_a = missing_same_day_a or set()
    missing_same_day_b_c = missing_same_day_b_c or set()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Planning"
    sheet["H13"] = "FORECAST"
    sheet["H14"] = 0.1
    sheet["H19"] = "FORECAST SAME DAY A"
    sheet["H20"] = "FORECAST SAME DAY B - C"
    sheet["G24"] = "Turno"
    sheet["H24"] = "drivers"
    start = date(2025, 12, 28)
    for offset in range(days):
        column = 12 + offset
        sheet.cell(row=24, column=column, value=start + timedelta(days=offset))
        if offset not in missing_next_day:
            sheet.cell(row=13, column=column, value=76 if offset % 2 == 0 else 78)
        if offset not in missing_same_day_a:
            sheet.cell(row=19, column=column, value=20)
        if offset not in missing_same_day_b_c:
            sheet.cell(row=20, column=column, value=18)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _legacy_import(
    content: bytes,
    *,
    organization_id: str = ORG,
    summary: dict | None = None,
) -> int:
    with db_session() as conn:
        cursor = conn.execute(
            """
            INSERT INTO workforce_imports (
                fingerprint, original_filename, imported_at, sheets,
                summary, organization_id
            ) VALUES (?, ?, ?, '[]', ?, ?)
            """,
            (
                sha256(content).hexdigest(),
                "Planning driver_DLO2_2026.xlsx",
                "2026-08-08T10:23:10Z",
                json.dumps(summary or {}),
                organization_id,
            ),
        )
        return int(cursor.lastrowid)


def _upload(content: bytes) -> dict:
    return {
        "file": (
            "Planning driver_DLO2_2026.xlsx",
            content,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }


def _preview(content: bytes, import_id: int):
    return client.post(
        f"{BASE}/preview",
        data={"workforce_import_id": str(import_id)},
        files=_upload(content),
    )


def _apply(content: bytes, import_id: int, preview_fingerprint: str):
    return client.post(
        BASE,
        data={
            "workforce_import_id": str(import_id),
            "expected_preview_fingerprint": preview_fingerprint,
        },
        files=_upload(content),
    )


def _coverage_rows(organization_id: str = ORG):
    with db_session() as conn:
        return conn.execute(
            """
            SELECT * FROM workforce_daily_coverage_requirements
            WHERE organization_id = ? ORDER BY operational_date, operational_cycle,
                 coverage_segment, id
            """,
            (organization_id,),
        ).fetchall()


def test_preview_requires_exact_recoverable_source_and_never_infers_forecast():
    content = _planning_workbook()
    import_id = _legacy_import(content)

    unavailable = client.post(
        f"{BASE}/preview", data={"workforce_import_id": str(import_id)}
    )
    assert unavailable.status_code == 200
    assert unavailable.json()["status"] == "SOURCE_NOT_RECOVERABLE"
    assert unavailable.json()["source_recoverable"] is False
    assert _coverage_rows() == []

    wrong_content = _planning_workbook(days=3)
    mismatch = _preview(wrong_content, import_id)
    assert mismatch.status_code == 200
    assert mismatch.json()["status"] == "SOURCE_MISMATCH"
    assert _coverage_rows() == []


def test_preview_and_apply_restore_only_exact_forecast_with_provenance():
    content = _planning_workbook(
        missing_same_day_a={1}, missing_same_day_b_c={1}
    )
    import_id = _legacy_import(content)
    preview = _preview(content, import_id)
    assert preview.status_code == 200, preview.text
    body = preview.json()
    assert body["status"] == "READY"
    assert body["requirements_expected"] == 4
    assert body["next_day_count"] == 2
    assert body["same_day_a_count"] == 1
    assert body["same_day_b_c_count"] == 1
    assert body["period_start"] == "2025-12-28"
    assert body["period_end"] == "2025-12-29"

    applied = _apply(content, import_id, body["preview_fingerprint"])
    assert applied.status_code == 200, applied.text
    result = applied.json()
    assert result["requirements_created"] == 4
    assert result["requirements_skipped"] == 0
    assert result["status"] == "ALREADY_COMPLETE"
    rows = _coverage_rows()
    assert len(rows) == 4
    assert {row["source"] for row in rows} == {"LEGACY_IMPORT_BACKFILL"}
    assert all(
        row["source_identity"].startswith(f"legacy-backfill:import:{import_id}:")
        for row in rows
    )
    assert {row["source_reference"] for row in rows} == {
        "Planning!L13", "Planning!M13", "Planning!L19", "Planning!L20"
    }
    next_day = next(
        row for row in rows
        if row["operational_date"] == "2025-12-28"
        and row["operational_cycle"] == "NEXT_DAY"
    )
    assert (next_day["forecast_routes"], next_day["required_capacity"]) == (76, 84)


def test_apply_is_idempotent_and_stale_preview_is_rejected():
    content = _planning_workbook()
    import_id = _legacy_import(content)
    preview = _preview(content, import_id).json()

    conflict = _apply(content, import_id, "0" * 64)
    assert conflict.status_code == 409
    assert _coverage_rows() == []

    first = _apply(content, import_id, preview["preview_fingerprint"])
    second = _apply(content, import_id, preview["preview_fingerprint"])
    assert first.status_code == 200
    assert second.status_code == 200
    assert first.json()["requirements_created"] == 6
    assert second.json()["requirements_created"] == 0
    assert second.json()["requirements_skipped"] == 6
    assert second.json()["idempotent"] is True
    assert len(_coverage_rows()) == 6


def test_existing_modern_requirement_is_never_overwritten():
    content = _planning_workbook(days=1)
    import_id = _legacy_import(content)
    modern = ImportedDailyCoverageRequirement(
        operational_date="2025-12-28",
        station=None,
        operational_cycle="NEXT_DAY",
        coverage_segment=None,
        forecast_routes=999,
        reserve_percentage=12,
        required_capacity=1119,
        source="IMPORT",
        source_reference="modern.xlsx!L13",
        source_identity="import:modern-source",
    )
    with db_session() as conn:
        coverage_repository.persist_imported_requirements(
            conn, [modern], organization_id=ORG, now=utc_now_iso()
        )

    preview = _preview(content, import_id).json()
    assert preview["existing_modern_rows"] == 1
    assert preview["requirements_missing"] == 2
    result = _apply(content, import_id, preview["preview_fingerprint"])
    assert result.status_code == 200, result.text
    assert result.json()["requirements_created"] == 2
    rows = _coverage_rows()
    preserved = next(
        row for row in rows if row["operational_cycle"] == "NEXT_DAY"
    )
    assert preserved["forecast_routes"] == 999
    assert preserved["source_identity"] == "import:modern-source"
    assert len(rows) == 3


def test_apply_is_atomic_and_rolls_back_on_insert_failure(monkeypatch):
    content = _planning_workbook()
    import_id = _legacy_import(content)
    preview = legacy_coverage_backfill_service.preview(
        ORG,
        content=content,
        filename="Planning driver_DLO2_2026.xlsx",
        workforce_import_id=import_id,
    )
    original = legacy_coverage_backfill_repository._insert_rows

    def fail_after_first(conn, rows):
        original(conn, rows[:1])
        raise RuntimeError("forced rollback")

    monkeypatch.setattr(
        legacy_coverage_backfill_repository, "_insert_rows", fail_after_first
    )
    with pytest.raises(RuntimeError, match="forced rollback"):
        legacy_coverage_backfill_service.apply(
            ORG,
            content=content,
            filename="Planning driver_DLO2_2026.xlsx",
            workforce_import_id=import_id,
            expected_preview_fingerprint=preview.preview_fingerprint or "",
        )
    assert _coverage_rows() == []


def test_backfill_is_organization_scoped_and_does_not_mutate_canonical_workforce():
    content = _planning_workbook()
    foreign_import_id = _legacy_import(content, organization_id="other-org")
    foreign = _preview(content, foreign_import_id)
    assert foreign.status_code == 200
    assert foreign.json()["status"] == "NO_ELIGIBLE_IMPORT"

    import_id = _legacy_import(content)
    with db_session() as conn:
        before = {
            "members": conn.execute(
                "SELECT COUNT(*) total FROM workforce_members"
            ).fetchone()["total"],
            "statuses": conn.execute(
                "SELECT COUNT(*) total FROM workforce_day_statuses"
            ).fetchone()["total"],
            "requirements": conn.execute(
                "SELECT COUNT(*) total FROM workforce_requirements"
            ).fetchone()["total"],
        }
    preview = _preview(content, import_id).json()
    assert _apply(content, import_id, preview["preview_fingerprint"]).status_code == 200
    with db_session() as conn:
        after = {
            "members": conn.execute(
                "SELECT COUNT(*) total FROM workforce_members"
            ).fetchone()["total"],
            "statuses": conn.execute(
                "SELECT COUNT(*) total FROM workforce_day_statuses"
            ).fetchone()["total"],
            "requirements": conn.execute(
                "SELECT COUNT(*) total FROM workforce_requirements"
            ).fetchone()["total"],
        }
    assert after == before
    assert _coverage_rows("other-org") == []


def test_modern_import_is_not_eligible_for_legacy_backfill():
    content = _planning_workbook()
    import_id = _legacy_import(
        content, summary={"coverage_requirements_detected": 6}
    )
    preview = _preview(content, import_id)
    assert preview.status_code == 200
    assert preview.json()["status"] == "NO_ELIGIBLE_IMPORT"
    assert _coverage_rows() == []


def test_real_period_fixture_restores_exact_1109_counts_and_missing_cells():
    content = _planning_workbook(
        days=372,
        missing_next_day={10, 11},
        missing_same_day_a={20, 21, 22},
        missing_same_day_b_c={30, 31},
    )
    import_id = _legacy_import(content)
    preview, requirements = legacy_coverage_backfill_service.inspect(
        ORG,
        content=content,
        filename="Planning driver_DLO2_2026.xlsx",
        workforce_import_id=import_id,
    )
    assert preview.period_start == "2025-12-28"
    assert preview.period_end == "2027-01-03"
    assert preview.next_day_count == 370
    assert preview.same_day_a_count == 369
    assert preview.same_day_b_c_count == 370
    assert preview.requirements_expected == 1109
    indexed = {
        (item.operational_date, item.operational_cycle, item.coverage_segment): item
        for item in requirements
    }
    assert indexed[("2025-12-28", "NEXT_DAY", None)].required_capacity == 84
    assert indexed[("2025-12-29", "NEXT_DAY", None)].required_capacity == 86
    assert ("2026-01-07", "NEXT_DAY", None) not in indexed
    assert ("2026-01-27", "SAME_DAY", "B_C") not in indexed


def test_backfilled_rows_are_visible_through_existing_coverage_read_model():
    content = _planning_workbook(
        days=1,
        missing_same_day_a={0},
        missing_same_day_b_c={0},
    )
    import_id = _legacy_import(content)
    with db_session() as conn:
        cursor = conn.execute(
            """
            INSERT INTO workforce_members (
                external_identifier, display_name, capabilities, active,
                source_reference, created_at, updated_at, organization_id,
                operational_cycle
            ) VALUES (
                'driver-1', 'Driver One', '[]', 1, 'legacy-planning',
                '2026-08-13T10:00:00Z', '2026-08-13T10:00:00Z', ?, 'NEXT_DAY'
            )
            """,
            (ORG,),
        )
        conn.execute(
            """
            INSERT INTO workforce_day_statuses (
                workforce_member_id, date, status_code, availability,
                shift_code, source_reference, observed_or_confirmed,
                updated_at, organization_id
            ) VALUES (?, '2025-12-28', 'scheduled', 1, 'C1',
                      'legacy-planning', 'imported', '2026-08-13T10:00:00Z', ?)
            """,
            (int(cursor.lastrowid), ORG),
        )
    preview = _preview(content, import_id).json()
    assert _apply(content, import_id, preview["preview_fingerprint"]).status_code == 200

    response = client.get(
        "/api/plugins/workforce/v1/planning/coverage",
        params={"date_from": "2025-12-28", "date_to": "2025-12-28"},
    )
    assert response.status_code == 200, response.text
    item = next(
        row for row in response.json()["items"] if row["cycle"] == "NEXT_DAY"
    )
    assert item["forecast_routes"] == 76
    assert item["required_capacity"] == 84
    assert item["assigned_drivers"] == 1
    assert item["requirement_gap"] == 83
    assert item["source"] == "LEGACY_IMPORT_BACKFILL"
