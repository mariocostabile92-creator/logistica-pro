import io
import json
import tempfile
from datetime import date, timedelta
from functools import lru_cache
from time import perf_counter

import pytest
from openpyxl import Workbook

from app.core.database import PostgresConnection, db_session
from app.plugins.workforce.application import workforce_service
from app.plugins.workforce.application.preview_cache import (
    WorkforcePreviewCache,
)
from app.plugins.workforce.domain.errors import (
    WorkforceImportConfirmationError,
)
from app.plugins.workforce.importer.workbook_interpreter import (
    interpret_workforce_workbook,
)
from app.plugins.workforce.infrastructure import import_repository


ACTIVE_DAYS = 273
MEMBER_COUNT = 173
STATUS_COUNT = MEMBER_COUNT * ACTIVE_DAYS


@lru_cache(maxsize=4)
def annual_workforce_book(
    member_count: int = MEMBER_COUNT,
    active_days: int = ACTIVE_DAYS,
    calendar_days: int = 365,
) -> bytes:
    book = Workbook()
    members = book.active
    members.title = "Anagrafiche e contratti"
    members.append(
        [
            "Matricola",
            "Nome Cognome",
            "Ruolo",
            "Tipo contratto",
            "Inizio contratto",
            "Fine contratto",
            "Ore settimanali",
            "Capability",
        ]
    )
    for index in range(member_count):
        members.append(
            [
                f"SYN-PERF-{index + 1:03d}",
                f"Risorsa Sintetica {index + 1:03d}",
                "courier",
                "part-time" if index % 3 == 0 else "full-time",
                "2026-01-01",
                "2026-12-31",
                24 if index % 3 == 0 else 40,
                "license_b,electric" if index % 5 == 0 else "license_b",
            ]
        )

    start = date(2026, 1, 1)
    days = [start + timedelta(days=offset) for offset in range(calendar_days)]
    schedule = book.create_sheet("Turni annuali")
    schedule.append(["Matricola", "Nome Cognome", *days])
    codes = ("M", "P", "S1", "R", "F")
    for index in range(member_count):
        populated = [
            codes[(index + day_index) % len(codes)]
            for day_index in range(active_days)
        ]
        schedule.append(
            [
                f"SYN-PERF-{index + 1:03d}",
                f"Risorsa Sintetica {index + 1:03d}",
                *populated,
                *([None] * (calendar_days - active_days)),
            ]
        )

    requirements = book.create_sheet("Fabbisogno")
    requirements.append(["Data", "Sede", "Fabbisogno", "Capability"])
    for day in days[:active_days]:
        requirements.append([day, "UNIT-SYN", 150, "license_b"])

    output = io.BytesIO()
    book.save(output)
    book.close()
    return output.getvalue()


@pytest.fixture(autouse=True)
def clear_workforce_preview_cache():
    workforce_service.preview_cache.clear()
    yield
    workforce_service.preview_cache.clear()


def table_count(table: str) -> int:
    with db_session() as conn:
        return int(conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0])


def test_annual_import_reuses_preview_and_meets_local_targets(monkeypatch):
    content = annual_workforce_book()
    parse_calls = 0
    captured_metrics = {}
    original_interpret = workforce_service.interpret_workforce_workbook
    original_persist = import_repository.apply_import

    def measured_interpret(*args, **kwargs):
        nonlocal parse_calls
        parse_calls += 1
        return original_interpret(*args, **kwargs)

    def measured_persist(*args, **kwargs):
        result = original_persist(*args, **kwargs)
        captured_metrics.update(kwargs["metrics"])
        return result

    monkeypatch.setattr(
        workforce_service,
        "interpret_workforce_workbook",
        measured_interpret,
    )
    monkeypatch.setattr(import_repository, "apply_import", measured_persist)

    analysis_started = perf_counter()
    preview = workforce_service.preview_import(
        content,
        "synthetic-annual.xlsx",
    )
    analysis_seconds = perf_counter() - analysis_started
    import_started = perf_counter()
    result = workforce_service.apply_import(
        content,
        "synthetic-annual.xlsx",
        preview.fingerprint,
        actor="performance-test",
    )
    import_seconds = perf_counter() - import_started

    assert analysis_seconds < 10
    assert import_seconds < 15
    assert analysis_seconds + import_seconds < 25
    assert parse_calls == 1
    assert preview.people_detected == MEMBER_COUNT
    assert result.statuses_created == STATUS_COUNT
    assert int(captured_metrics["database_calls"]) <= 60
    assert int(captured_metrics["bulk_batches"]) == 52
    assert table_count("workforce_members") == MEMBER_COUNT
    assert table_count("workforce_day_statuses") == STATUS_COUNT
    assert table_count("workforce_changes") == STATUS_COUNT + MEMBER_COUNT
    assert table_count("workforce_import_rows") == (
        STATUS_COUNT + (MEMBER_COUNT * 2)
    )

    response = result.model_dump_json()
    response_data = json.loads(response)
    assert len(response.encode("utf-8")) < 1024
    assert "matrix" not in response
    assert set(response_data) == {
        "fingerprint",
        "idempotent",
        "members_created",
        "members_updated",
        "statuses_created",
        "statuses_updated",
        "requirements_created",
        "coverage_requirements_created",
        "coverage_requirements_updated",
        "sheets_imported",
    }


def test_second_import_is_idempotent_without_parsing_again(monkeypatch):
    content = annual_workforce_book(member_count=3, active_days=5, calendar_days=7)
    preview = workforce_service.preview_import(content, "small-annual.xlsx")
    first = workforce_service.apply_import(
        content,
        "small-annual.xlsx",
        preview.fingerprint,
    )

    def fail_if_reparsed(*_args, **_kwargs):
        raise AssertionError("An imported fingerprint must not be parsed again.")

    monkeypatch.setattr(
        workforce_service,
        "interpret_workforce_workbook",
        fail_if_reparsed,
    )
    second = workforce_service.apply_import(
        content,
        "small-annual.xlsx",
        preview.fingerprint,
    )

    assert first.idempotent is False
    assert second.idempotent is True
    assert table_count("workforce_day_statuses") == 15


def test_expired_preview_falls_back_to_safe_reparse(monkeypatch):
    clock = [100.0]
    cache = WorkforcePreviewCache(
        ttl_seconds=1,
        max_entries=2,
        clock=lambda: clock[0],
    )
    monkeypatch.setattr(workforce_service, "preview_cache", cache)
    content = annual_workforce_book(member_count=2, active_days=2, calendar_days=2)
    calls = 0
    original = workforce_service.interpret_workforce_workbook

    def counted(*args, **kwargs):
        nonlocal calls
        calls += 1
        return original(*args, **kwargs)

    monkeypatch.setattr(
        workforce_service,
        "interpret_workforce_workbook",
        counted,
    )
    preview = workforce_service.preview_import(content, "expired.xlsx")
    clock[0] += 2

    result = workforce_service.apply_import(
        content,
        "expired.xlsx",
        preview.fingerprint,
    )

    assert calls == 2
    assert result.statuses_created == 4


def test_changed_fingerprint_is_rejected_before_reparse(monkeypatch):
    content = annual_workforce_book(member_count=2, active_days=2, calendar_days=2)
    preview = workforce_service.preview_import(content, "fingerprint.xlsx")

    def fail_if_called(*_args, **_kwargs):
        raise AssertionError("Mismatch must fail before workbook parsing.")

    monkeypatch.setattr(
        workforce_service,
        "interpret_workforce_workbook",
        fail_if_called,
    )
    with pytest.raises(WorkforceImportConfirmationError):
        workforce_service.apply_import(
            content + b"changed",
            "fingerprint.xlsx",
            preview.fingerprint,
        )


def test_bulk_failure_rolls_back_members_statuses_audit_and_import(monkeypatch):
    content = annual_workforce_book(member_count=3, active_days=4, calendar_days=4)
    parsed = interpret_workforce_workbook(content, "atomic.xlsx")
    original = import_repository._executemany

    def fail_after_status_insert(conn, metrics, statement, rows, chunk_size):
        original(conn, metrics, statement, rows, chunk_size)
        if "INSERT INTO workforce_day_statuses" in statement:
            raise RuntimeError("Synthetic bulk failure")

    monkeypatch.setattr(
        import_repository,
        "_executemany",
        fail_after_status_insert,
    )
    with pytest.raises(RuntimeError, match="Synthetic bulk failure"):
        import_repository.apply_import(
            parsed,
            original_filename="atomic.xlsx",
            actor="test",
            chunk_size=2,
        )

    for table in (
        "workforce_imports",
        "workforce_members",
        "workforce_day_statuses",
        "workforce_daily_coverage_requirements",
        "workforce_changes",
    ):
        assert table_count(table) == 0


@pytest.mark.parametrize(
    ("size", "expected"),
    ((500, [500, 500, 200]), (1000, [1000, 200]), (2000, [1200]), (5000, [1200])),
)
def test_chunking_preserves_all_rows(size, expected):
    batches = [len(batch) for batch in import_repository._chunks(list(range(1200)), size)]

    assert batches == expected
    assert sum(batches) == 1200


def test_preview_cache_is_bounded_and_does_not_store_file_bytes():
    parsed = interpret_workforce_workbook(
        annual_workforce_book(member_count=1, active_days=1, calendar_days=1),
        "privacy.xlsx",
    )
    cache = WorkforcePreviewCache(ttl_seconds=60, max_entries=1)
    cache.store(parsed)
    cached = cache.get(parsed.fingerprint)

    assert cached is parsed
    assert not hasattr(cached, "content")
    assert not hasattr(cached, "file_bytes")
    assert not hasattr(cached, "original_filename")


def test_import_does_not_leave_temporary_files(monkeypatch, tmp_path):
    monkeypatch.setenv("TEMP", str(tmp_path))
    monkeypatch.setenv("TMP", str(tmp_path))
    monkeypatch.setenv("TMPDIR", str(tmp_path))
    monkeypatch.setattr(tempfile, "tempdir", None)
    content = annual_workforce_book(member_count=2, active_days=2, calendar_days=2)

    preview = workforce_service.preview_import(content, "no-temporary.xlsx")
    workforce_service.apply_import(
        content,
        "no-temporary.xlsx",
        preview.fingerprint,
    )

    assert list(tmp_path.iterdir()) == []


def test_postgres_bulk_execution_translates_placeholders_without_returning():
    calls = []

    class Cursor:
        description = None

        def executemany(self, statement, parameters):
            calls.append((statement, parameters))

    class Connection:
        def cursor(self):
            return Cursor()

    connection = PostgresConnection(Connection())
    connection.executemany(
        "INSERT INTO workforce_members (external_identifier) VALUES (?)",
        [("SYN-1",), ("SYN-2",)],
    )

    assert calls[0][0].endswith("VALUES (%s)")
    assert "RETURNING" not in calls[0][0]
    assert len(calls[0][1]) == 2
