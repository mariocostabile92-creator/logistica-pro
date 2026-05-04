from fastapi import FastAPI, HTTPException, UploadFile, File, Header
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from optimizer import optimize_route
from openpyxl import load_workbook

import io
import uuid
import hashlib
import secrets

from database import (
    init_db,
    get_user_by_email,
    get_user_by_token,
    create_user,
    update_user_token,
    update_user_password,
    save_route_history,
    get_route_history,
    delete_route_history,
    save_delivery_report,
    get_delivery_reports,
    delete_delivery_reports,
    get_user_analytics,
)


app = FastAPI(
    title="Logistica Pro MVP",
    version="3.0.0"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def startup():
    init_db()


class RegisterRequest(BaseModel):
    email: str
    password: str


class LoginRequest(BaseModel):
    email: str
    password: str


class ResetPasswordRequest(BaseModel):
    email: str
    recovery_code: str
    new_password: str


class RouteRequest(BaseModel):
    depot: str
    deliveries: list[str]


def hash_password(password: str):
    return hashlib.sha256(password.encode("utf-8")).hexdigest()


def create_token():
    return secrets.token_hex(32)


def create_recovery_code():
    return str(uuid.uuid4())[:8].upper()


def get_current_user(authorization: str | None):
    if not authorization:
        raise HTTPException(status_code=401, detail="Token mancante. Effettua il login.")

    token = authorization.replace("Bearer ", "").strip()
    user = get_user_by_token(token)

    if not user:
        raise HTTPException(status_code=401, detail="Token non valido. Effettua di nuovo il login.")

    return user


@app.get("/")
def home():
    return {
        "message": "Logistica Pro MVP API collegata a PostgreSQL"
    }


@app.post("/register")
def register(request: RegisterRequest):
    email = request.email.lower().strip()
    password = request.password.strip()

    if not email or "@" not in email:
        raise HTTPException(status_code=400, detail="Email non valida.")

    if len(password) < 6:
        raise HTTPException(status_code=400, detail="La password deve avere almeno 6 caratteri.")

    existing_user = get_user_by_email(email)

    if existing_user:
        raise HTTPException(status_code=400, detail="Email già registrata.")

    user_id = str(uuid.uuid4())
    recovery_code = create_recovery_code()
    token = create_token()

    create_user(
        user_id=user_id,
        email=email,
        password_hash=hash_password(password),
        recovery_code=recovery_code,
        token=token
    )

    return {
        "message": "Utente registrato correttamente.",
        "token": token,
        "email": email,
        "recovery_code": recovery_code,
        "important": "Conserva questo codice recupero password."
    }


@app.post("/login")
def login(request: LoginRequest):
    email = request.email.lower().strip()
    password_hash = hash_password(request.password.strip())

    user = get_user_by_email(email)

    if not user:
        raise HTTPException(status_code=401, detail="Email o password non corretti.")

    if user["password_hash"] != password_hash:
        raise HTTPException(status_code=401, detail="Email o password non corretti.")

    token = create_token()
    update_user_token(email, token)

    return {
        "message": "Login effettuato.",
        "token": token,
        "email": email
    }


@app.post("/reset-password")
def reset_password(request: ResetPasswordRequest):
    email = request.email.lower().strip()
    recovery_code = request.recovery_code.upper().strip()
    new_password = request.new_password.strip()

    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="La nuova password deve avere almeno 6 caratteri.")

    user = get_user_by_email(email)

    if not user:
        raise HTTPException(status_code=400, detail="Email o codice recupero non validi.")

    if user["recovery_code"] != recovery_code:
        raise HTTPException(status_code=400, detail="Email o codice recupero non validi.")

    token = create_token()
    new_recovery_code = create_recovery_code()

    update_user_password(
        email=email,
        password_hash=hash_password(new_password),
        recovery_code=new_recovery_code,
        token=token
    )

    return {
        "message": "Password aggiornata correttamente.",
        "token": token,
        "new_recovery_code": new_recovery_code,
        "important": "Conserva il nuovo codice recupero password."
    }


@app.get("/me")
def me(authorization: str | None = Header(default=None)):
    user = get_current_user(authorization)

    return {
        "id": user["id"],
        "email": user["email"]
    }


def normalizza_testo(value):
    if value is None:
        return ""
    return str(value).strip()


def is_header_row(text: str):
    text = text.lower().strip()

    parole_header = [
        "cliente",
        "indirizzo",
        "telefono",
        "note",
        "address",
        "phone",
        "customer",
        "nome",
        "destinazione"
    ]

    return any(word in text for word in parole_header)


def sembra_indirizzo(value: str):
    value_lower = value.lower().strip()

    if len(value_lower) < 4:
        return False

    if value_lower.isdigit():
        return False

    parole_indirizzo = [
        "via ",
        "viale ",
        "piazza ",
        "corso ",
        "largo ",
        "strada ",
        "vicolo ",
        "loc.",
        "località"
    ]

    if "," in value_lower:
        return True

    return any(p in value_lower for p in parole_indirizzo)


def crea_delivery(cliente="", indirizzo="", telefono="", note=""):
    return {
        "cliente": normalizza_testo(cliente),
        "indirizzo": normalizza_testo(indirizzo),
        "telefono": normalizza_testo(telefono),
        "note": normalizza_testo(note)
    }


def aggiungi_delivery(deliveries: list, delivery: dict):
    indirizzo = delivery.get("indirizzo", "").strip()

    if not indirizzo:
        return

    if not sembra_indirizzo(indirizzo):
        return

    indirizzi_esistenti = [
        d["indirizzo"].strip().lower()
        for d in deliveries
    ]

    if indirizzo.strip().lower() not in indirizzi_esistenti:
        deliveries.append(delivery)


@app.post("/optimize-route")
def optimize_route_api(
    request: RouteRequest,
    authorization: str | None = Header(default=None)
):
    user = get_current_user(authorization)

    try:
        depot = request.depot.strip()

        if not depot:
            raise ValueError("Inserisci l'indirizzo del deposito.")

        deliveries_clean = []

        parole_da_ignorare = [
            "indirizzo",
            "indirizzi",
            "address",
            "addresses",
            "consegna",
            "consegne",
            "cliente",
            "telefono",
            "note"
        ]

        for item in request.deliveries:
            value = str(item).strip()

            if not value:
                continue

            if value.lower() in parole_da_ignorare:
                continue

            if value not in deliveries_clean:
                deliveries_clean.append(value)

        if not deliveries_clean:
            raise ValueError("Inserisci almeno una consegna valida.")

        result = optimize_route(
            depot_address=depot,
            delivery_addresses=deliveries_clean
        )

        save_route_history(
            user_id=user["id"],
            depot=depot,
            deliveries=deliveries_clean,
            result=result
        )

        return result

    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/import-excel")
async def import_excel(
    file: UploadFile = File(...),
    authorization: str | None = Header(default=None)
):
    get_current_user(authorization)

    try:
        content = await file.read()
        workbook = load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.active

        deliveries = []
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            return {
                "deliveries": [],
                "addresses": [],
                "count": 0
            }

        headers = [
            normalizza_testo(cell).lower()
            for cell in rows[0]
        ]

        has_headers = any(h in headers for h in [
            "cliente",
            "indirizzo",
            "telefono",
            "note",
            "address",
            "phone",
            "customer"
        ])

        if has_headers:
            for row in rows[1:]:
                row_data = {}

                for i, cell in enumerate(row):
                    if i >= len(headers):
                        continue

                    key = headers[i]
                    value = normalizza_testo(cell)
                    row_data[key] = value

                indirizzo = (
                    row_data.get("indirizzo", "")
                    or row_data.get("address", "")
                    or row_data.get("destinazione", "")
                )

                cliente = (
                    row_data.get("cliente", "")
                    or row_data.get("customer", "")
                    or row_data.get("nome", "")
                )

                telefono = (
                    row_data.get("telefono", "")
                    or row_data.get("phone", "")
                    or row_data.get("cellulare", "")
                )

                note = (
                    row_data.get("note", "")
                    or row_data.get("nota", "")
                    or row_data.get("notes", "")
                )

                delivery = crea_delivery(
                    cliente=cliente,
                    indirizzo=indirizzo,
                    telefono=telefono,
                    note=note
                )

                aggiungi_delivery(deliveries, delivery)

        for row in rows:
            for cell in row:
                value = normalizza_testo(cell)

                if not value:
                    continue

                if "|" not in value:
                    continue

                if is_header_row(value):
                    continue

                parts = [
                    p.strip()
                    for p in value.split("|")
                    if p.strip()
                ]

                if len(parts) >= 2:
                    delivery = crea_delivery(
                        cliente=parts[0],
                        indirizzo=parts[1],
                        telefono=parts[2] if len(parts) >= 3 else "",
                        note=parts[3] if len(parts) >= 4 else ""
                    )

                    aggiungi_delivery(deliveries, delivery)

        if len(deliveries) == 0:
            for row in rows:
                for cell in row:
                    value = normalizza_testo(cell)

                    if not value:
                        continue

                    if is_header_row(value):
                        continue

                    if sembra_indirizzo(value):
                        delivery = crea_delivery(
                            cliente="",
                            indirizzo=value,
                            telefono="",
                            note=""
                        )

                        aggiungi_delivery(deliveries, delivery)

        return {
            "deliveries": deliveries,
            "addresses": [d["indirizzo"] for d in deliveries],
            "count": len(deliveries)
        }

    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Errore import Excel: {str(e)}"
        )


@app.get("/storico")
def get_storico(authorization: str | None = Header(default=None)):
    user = get_current_user(authorization)

    storico = get_route_history(user["id"])

    return storico


@app.delete("/storico")
def delete_storico(authorization: str | None = Header(default=None)):
    user = get_current_user(authorization)

    delete_route_history(user["id"])

    return {
        "message": "Storico cancellato correttamente"
    }


@app.post("/save-report")
def save_report(
    payload: dict,
    authorization: str | None = Header(default=None)
):
    user = get_current_user(authorization)

    save_delivery_report(
        user_id=user["id"],
        payload=payload
    )

    return {
        "message": "Report consegne salvato correttamente"
    }


@app.get("/reports")
def get_reports(authorization: str | None = Header(default=None)):
    user = get_current_user(authorization)

    reports = get_delivery_reports(user["id"])

    return reports


@app.delete("/reports")
def delete_reports(authorization: str | None = Header(default=None)):
    user = get_current_user(authorization)

    delete_delivery_reports(user["id"])

    return {
        "message": "Report consegne cancellati correttamente"
    }


@app.get("/analytics")
def get_analytics(authorization: str | None = Header(default=None)):
    user = get_current_user(authorization)

    return get_user_analytics(user["id"])